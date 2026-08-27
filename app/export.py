"""Multi-sheet Excel export."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .db import conn
from .validate import pieces


def _header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1e293b")
    cell.alignment = Alignment(horizontal="center")


def export_bills() -> bytes:
    """Bills workbook: bills summary, all items, suppliers summary."""
    wb = Workbook()

    # Sheet 1: Bills summary
    ws1 = wb.active
    ws1.title = "Bills"
    headers = ["Bill ID", "Supplier", "Phone", "Date", "Bill No", "Status",
               "Payment", "Written Total", "Computed Total", "Unit", "Flags"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        _header_style(cell)

    with conn() as c:
        rows = c.execute(
            "SELECT * FROM bills WHERE deleted_at IS NULL ORDER BY id"
        ).fetchall()
        for i, r in enumerate(rows, 2):
            import json
            flags = json.loads(r["flags"] or "[]")
            vals = [r["id"], r["supplier_name"], r["phone"], r["bill_date"], r["bill_no"],
                    r["status"], r["payment_status"], r["written_total"], r["computed_total"],
                    r["unit"], "; ".join(flags)]
            for col, v in enumerate(vals, 1):
                ws1.cell(row=i, column=col, value=v)
    ws1.freeze_panes = "A2"

    # Sheet 2: All items detail
    ws2 = wb.create_sheet("Items")
    headers = ["Bill ID", "Supplier", "Date", "Raw", "Item Code", "Price", "Qty",
               "Unit", "Pieces", "Line Total", "Category", "Payment Status"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        _header_style(cell)

    with conn() as c:
        rows = c.execute(
            "SELECT bi.*, b.supplier_name, b.bill_date, b.payment_status, "
            "pc.name AS cat_name FROM bill_items bi "
            "JOIN bills b ON bi.bill_id = b.id "
            "AND b.deleted_at IS NULL "
            "LEFT JOIN price_categories pc ON bi.category_id = pc.id "
            "ORDER BY b.id, bi.id"
        ).fetchall()
        for i, r in enumerate(rows, 2):
            vals = [r["bill_id"], r["supplier_name"], r["bill_date"], r["raw"],
                    r["item_code"], r["price"], r["qty"], r["unit"],
                    pieces(r["qty"], r["unit"]), r["line_total"], r["cat_name"],
                    r["payment_status"]]
            for col, v in enumerate(vals, 1):
                ws2.cell(row=i, column=col, value=v)
    ws2.freeze_panes = "A2"

    # Sheet 3: Suppliers summary
    ws3 = wb.create_sheet("Suppliers")
    headers = ["ID", "Name", "Phone", "Address", "Notes", "Bill Count",
               "Total Spent", "Outstanding", "Last Purchase"]
    for col, h in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        _header_style(cell)

    with conn() as c:
        rows = c.execute(
            "SELECT s.*, COUNT(b.id) AS bill_count, "
            "COALESCE(SUM(CASE WHEN b.written_total IS NOT NULL THEN b.written_total "
            "ELSE b.computed_total END), 0) AS total_spent, "
            "SUM(CASE WHEN b.payment_status='credit' THEN COALESCE(b.written_total, b.computed_total, 0) "
            "ELSE 0 END) AS outstanding, "
            "MAX(b.bill_date) AS last_purchase "
            "FROM suppliers s LEFT JOIN bills b ON s.id = b.supplier_id AND b.status='confirmed' AND b.deleted_at IS NULL "
            "WHERE s.deleted_at IS NULL "
            "GROUP BY s.id ORDER BY total_spent DESC"
        ).fetchall()
        for i, r in enumerate(rows, 2):
            vals = [r["id"], r["name"], r["phone"], r["address"], r["notes"],
                    r["bill_count"], round(r["total_spent"], 2), round(r["outstanding"], 2),
                    r["last_purchase"]]
            for col, v in enumerate(vals, 1):
                ws3.cell(row=i, column=col, value=v)
    ws3.freeze_panes = "A2"

    # Auto-size columns
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_insights() -> bytes:
    """Insights workbook: ABC, dead stock, price comparison, suppliers, monthly trends.
    Empty sheets are skipped. If all empty, returns a single 'No Data' sheet."""
    from .insights import abc_analysis, dead_stock, price_comparison
    from .reports import supplier_ranking

    wb = Workbook()
    has_any_data = False

    # Sheet 1: ABC Analysis
    abc = abc_analysis()
    if abc["items"]:
        has_any_data = True
        ws = wb.active
        ws.title = "ABC Analysis"
        headers = ["Item", "Class", "Total Qty", "Bill Count", "Revenue", "Cost", "Profit", "Margin"]
        for col, h in enumerate(headers, 1):
            _header_style(ws.cell(row=1, column=col, value=h))
        for i, item in enumerate(abc["items"], 2):
            for col, v in enumerate(
                [item["raw"], item["class"], item["total_qty"], item["bill_count"],
                 item["revenue"], item["cost"], item["profit"], f"{item['margin']*100:.1f}%"], 1):
                ws.cell(row=i, column=col, value=v)
    else:
        # Remove the default empty sheet — we'll create sheets only for non-empty data
        if wb.active and wb.active.title == "Sheet":
            wb.remove(wb.active)

    # Sheet 2: Dead Stock (skip if empty)
    ds = dead_stock(60)
    if ds["items"]:
        has_any_data = True
        ws2 = wb.create_sheet("Dead Stock")
        headers = ["Item", "Last Seen", "Total Qty", "Tied Capital"]
        for col, h in enumerate(headers, 1):
            _header_style(ws2.cell(row=1, column=col, value=h))
        for i, item in enumerate(ds["items"], 2):
            for col, v in enumerate(
                [item["raw"], item["last_seen"], item["total_qty"], item["tied_capital"]], 1):
                ws2.cell(row=i, column=col, value=v)

    # Sheet 3: Price Comparison (skip if empty)
    pc = price_comparison()
    if pc["items"]:
        has_any_data = True
        ws3 = wb.create_sheet("Price Comparison")
        headers = ["Item", "Best Supplier", "Best Price", "Worst Supplier", "Worst Price",
                   "Savings", "Savings %"]
        for col, h in enumerate(headers, 1):
            _header_style(ws3.cell(row=1, column=col, value=h))
        for i, item in enumerate(pc["items"], 2):
            for col, v in enumerate(
                [item["raw"], item["best_supplier"], item["best_price"],
                 item["worst_supplier"], item["worst_price"],
                 item["savings"], f"{item['savings_pct']*100:.1f}%"], 1):
                ws3.cell(row=i, column=col, value=v)

    # Sheet 4: Suppliers (skip if empty)
    sr = supplier_ranking()
    if sr["suppliers"]:
        has_any_data = True
        ws4 = wb.create_sheet("Suppliers")
        headers = ["Rank", "Supplier", "Phone", "Bills", "Total Spent", "Outstanding", "Last Purchase"]
        for col, h in enumerate(headers, 1):
            _header_style(ws4.cell(row=1, column=col, value=h))
        for i, s in enumerate(sr["suppliers"], 2):
            for col, v in enumerate(
                [i - 1, s["name"], s["phone"], s["bill_count"],
                 s["total_spent"], s["outstanding"], s["last_purchase"]], 1):
                ws4.cell(row=i, column=col, value=v)

    # Sheet 5: Monthly Trends (skip if empty)
    with conn() as c:
        rows = c.execute(
            "SELECT strftime('%Y-%m', bill_date) AS month, supplier_name, "
            "SUM(COALESCE(written_total, computed_total)) AS spend "
            "FROM bills WHERE status='confirmed' AND deleted_at IS NULL AND bill_date IS NOT NULL "
            "GROUP BY month, supplier_name ORDER BY month, supplier_name"
        ).fetchall()
    if rows:
        has_any_data = True
        ws5 = wb.create_sheet("Monthly Trends")
        headers = ["Month", "Supplier", "Spend"]
        for col, h in enumerate(headers, 1):
            _header_style(ws5.cell(row=1, column=col, value=h))
        for i, r in enumerate(rows, 2):
            for col, v in enumerate([r["month"], r["supplier_name"], round(r["spend"], 2)], 1):
                ws5.cell(row=i, column=col, value=v)

    # If no data at all, create a single "No Data" sheet
    if not has_any_data:
        ws = wb.create_sheet("No Data")
        ws.cell(row=1, column=1, value="No data available for export.")

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
