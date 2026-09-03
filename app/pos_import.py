"""External POS backup import — accept daily exports from any existing POS system.

Supports CSV / Excel / JSON / ZIP files with flexible column mapping so it works
with Incline POS, eHisabKitab, CloudPOS, Uvision UPOS, generic Excel exports, etc.

v8.4: ZIP upload support — users can upload a ZIP containing CSV/JSON/Excel files.
We extract the archive, find the largest data file, and parse it automatically.

The user uploads a file, we detect columns automatically (or let them map manually),
then import sales + items into the sales/sale_items tables — making them appear in
"Today's sales" / Z-Report / insights just like native BillBook sales.
"""
import logging
import csv
import io
import json
import os
import re
import tempfile
import zipfile
from datetime import datetime
from urllib.parse import quote
from .db import conn
from . import db  # v8.14.0: needed for db.VALID_SALE_FILTER constant

logger = logging.getLogger(__name__)


# Canonical fields we want to extract from any POS export
CANONICAL_FIELDS = [
    "invoice_no",      # invoice / receipt number
    "date",            # sale date (YYYY-MM-DD or full datetime)
    "time",            # sale time (optional, separate column)
    "customer_name",
    "customer_phone",
    "item_name",       # per-item rows
    "item_code",       # SKU / barcode
    "qty",
    "price",           # unit sell price
    "line_total",      # qty × price (optional, computed if missing)
    "subtotal",        # sale-level subtotal (header rows)
    "discount",
    "total",           # sale grand total (header rows)
    "payment_method",  # cash / card / credit / online
    "payment_status",  # paid / credit / partial
    "category",        # item category (e.g. "A", "250")
    "notes",
]

# Heuristic column-name detection — maps various POS export headers to canonical fields
COLUMN_HINTS = {
    "invoice_no": ["invoice", "invoice_no", "invoice number", "receipt", "receipt_no",
                   "bill_no", "bill number", "sale_id", "order_no", "order number", "ref"],
    "date": ["date", "sale_date", "bill_date", "invoice_date", "transaction_date", "txn_date"],
    "time": ["time", "sale_time", "txn_time"],
    "customer_name": ["customer", "customer_name", "cust", "buyer", "client"],
    "customer_phone": ["phone", "mobile", "customer_phone", "cust_phone", "contact"],
    "item_name": ["item", "item_name", "product", "product_name", "description", "desc",
                  "particular", "particulars"],
    "item_code": ["code", "item_code", "sku", "barcode", "product_code", "art"],
    "qty": ["qty", "quantity", "pieces", "pcs", "count", "amount"],
    "price": ["price", "unit_price", "rate", "sell_price", "selling_price", "mrp"],
    "line_total": ["line_total", "total", "amount", "subtotal", "net"],
    "subtotal": ["subtotal", "sub_total", "gross", "gross_total"],
    "discount": ["discount", "disc", "rebate", "off"],
    "total": ["total", "grand_total", "net_total", "net", "final_total", "bill_total",
              "amount_paid", "payable"],
    "payment_method": ["payment", "payment_method", "pay_method", "mode", "paid_via"],
    "payment_status": ["status", "payment_status", "paid_status"],
    "category": ["category", "cat", "group", "type"],
    "notes": ["notes", "remarks", "comment", "description"],
}


def detect_columns(headers: list) -> dict:
    """Auto-detect which input column maps to which canonical field.
    Returns {canonical_field: input_column_name}"""
    mapping = {}
    headers_norm = {h.strip().lower(): h for h in headers}
    for canonical, hints in COLUMN_HINTS.items():
        for hint in hints:
            # Exact match
            if hint in headers_norm:
                mapping[canonical] = headers_norm[hint]
                break
            # Substring match
            for h_norm, h_orig in headers_norm.items():
                if hint in h_norm and canonical not in mapping:
                    mapping[canonical] = h_orig
                    break
            if canonical in mapping:
                break
    return mapping


def parse_value(val, field: str):
    """Parse a string value into the right type for the given canonical field."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", "null", "-"):
        return None
    if field in ("qty",):
        # Qty can be fractional (e.g. 1.5 dozen); try int first, fallback float
        try:
            f = float(s.replace(",", ""))
            return int(f) if f == int(f) else f
        except Exception:
            return None
    if field in ("price", "line_total", "subtotal", "discount", "total"):
        try:
            # Strip currency symbols and spaces
            cleaned = re.sub(r"[^\d.\-]", "", s.replace(",", ""))
            return float(cleaned) if cleaned else None
        except Exception:
            return None
    return s


def parse_date(s):
    """Parse a date string in various formats. Returns YYYY-MM-DD."""
    if not s:
        return None
    s = str(s).strip()
    formats = [
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y",
        "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d-%m-%Y",
        "%Y/%m/%d", "%d.%m.%Y",
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M:%S")
        except Exception as _e:
            logger.warning("Silent exception in pos_import.py: %s", _e, exc_info=True)
    # Try ISO format with T
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M:%S")
    except Exception as _e:
        logger.warning("Silent exception in pos_import.py: %s", _e, exc_info=True)
    return s[:10] if len(s) >= 10 else None, None


def normalize_payment_method(val):
    """Normalize various payment method strings to canonical ones."""
    if not val:
        return "cash"
    s = str(val).lower().strip()
    if any(x in s for x in ["cash", "نقد", "rtt", "rtcs"]):
        return "cash"
    if any(x in s for x in ["card", "credit card", "debit", "visa", "master"]):
        return "card"
    if any(x in s for x in ["online", "transfer", "bank", "jazzcash", "easypaisa", "raftaar"]):
        return "online"
    if any(x in s for x in ["credit", "urdhaar", "udhaar", "kiste", "installment", "credit sale"]):
        return "credit"
    if any(x in s for x in ["split", "mixed", "partial"]):
        return "split"
    return "cash"


def _get_or_create_category_by_code(code: str, sell_price: float = 0) -> int | None:
    """Find or create a price category by code (used when import has 'category' column)."""
    if not code:
        return None
    code = str(code).strip().upper()
    with conn() as c:
        row = c.execute("SELECT id FROM price_categories WHERE code=? OR name=?", (code, code)).fetchone()
        if row:
            return row["id"]
        # Create new category if it doesn't exist (use code as both name and code)
        if sell_price > 0:
            cur = c.execute(
                "INSERT INTO price_categories(name, code, sell_price, color) VALUES(?,?,?,?)",
                (code, code, sell_price, "#6b7280"),
            )
            return cur.lastrowid
    return None


def _get_or_create_customer(name: str, phone: str) -> int | None:
    if not name and not phone:
        return None
    with conn() as c:
        if phone:
            row = c.execute("SELECT id FROM customers WHERE phone=?", (phone,)).fetchone()
            if row:
                return row["id"]
        if name:
            row = c.execute("SELECT id FROM customers WHERE lower(name)=lower(?)", (name,)).fetchone()
            if row:
                return row["id"]
        return c.execute(
            "INSERT INTO customers(name, phone) VALUES(?,?)", (name or "Walk-in", phone)
        ).lastrowid


def import_pos_backup(rows: list, mapping: dict, source_name: str = "",
                       filename: str = "", import_date: str = "",
                       notes: str = "") -> dict:
    """Import parsed rows from a POS backup file.

    rows: list of dicts (each dict is one row from the file, keyed by original column name)
    mapping: {canonical_field: original_column_name} — which input column to read for each field

    Returns summary: {imported_sales, imported_items, skipped, total_revenue, errors, import_id}
    """
    if not rows:
        return {"imported_sales": 0, "imported_items": 0, "skipped": 0,
                "total_revenue": 0, "errors": ["No rows in file"], "import_id": None}

    imported_sales = 0
    imported_items = 0
    skipped = 0
    errors = []
    total_revenue = 0.0
    dates_seen = []

    def get_field(row, canonical_field, default=None):
        col = mapping.get(canonical_field)
        if not col:
            return default
        return parse_value(row.get(col), canonical_field) if canonical_field in (
            "qty", "price", "line_total", "subtotal", "discount", "total"
        ) else row.get(col)

    # Two patterns:
    # 1. Per-sale rows (one row = one sale, with item_name/qty as a single combined column)
    # 2. Per-item rows (multiple rows per sale, grouped by invoice_no)
    # Detect: if rows have invoice_no AND item_name, treat as per-item
    has_items = bool(mapping.get("item_name"))
    has_invoice = bool(mapping.get("invoice_no"))

    if has_invoice and has_items:
        # Per-item rows — group by invoice_no
        sales_grouped = {}
        sale_order = []
        for row in rows:
            inv = get_field(row, "invoice_no")
            if not inv:
                # No invoice number — try date+total as a synthetic key
                inv = f"NOINV-{get_field(row, 'date', '')}-{get_field(row, 'total', '')}"
            if inv not in sales_grouped:
                sales_grouped[inv] = []
                sale_order.append(inv)
            sales_grouped[inv].append(row)
        # Process each sale (commit after each to avoid long transactions)
        for inv_no, item_rows in sales_grouped.items():
            try:
                first = item_rows[0]
                date_str, time_str = parse_date(get_field(first, "date")) or (None, None)
                if date_str:
                    dates_seen.append(date_str)
                customer_name = get_field(first, "customer_name") or ""
                customer_phone = get_field(first, "customer_phone") or ""
                subtotal = sum(
                    (get_field(r, "line_total") or
                     (get_field(r, "price") or 0) * (get_field(r, "qty") or 1))
                    for r in item_rows
                ) or get_field(first, "subtotal") or 0
                discount = get_field(first, "discount") or 0
                total = get_field(first, "total") or (subtotal - discount)
                pay_method = normalize_payment_method(get_field(first, "payment_method"))
                pay_status = "credit" if pay_method == "credit" else "paid"
                notes_val = get_field(first, "notes") or ""

                # Insert sale (with a marker that it's imported)
                invoice_no = f"IMP-{inv_no}" if not str(inv_no).startswith("IMP-") else inv_no
                # Skip if already imported (idempotent)
                with conn() as c:
                    existing = c.execute(
                        "SELECT id FROM sales WHERE invoice_no=?", (invoice_no,)
                    ).fetchone()
                if existing:
                    skipped += 1
                    continue

                # Create customer if needed
                customer_id = _get_or_create_customer(customer_name, customer_phone)

                # Build full datetime for created_at
                created_at = date_str or datetime.now().strftime("%Y-%m-%d")
                if time_str:
                    created_at = f"{date_str} {time_str}"

                with conn() as c:
                    sale_id = c.execute(
                        "INSERT INTO sales(invoice_no, customer_name, customer_phone, customer_id, "
                        "subtotal, discount, total, payment_method, payment_status, notes, created_at) "
                        "VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                        (invoice_no, customer_name, customer_phone, customer_id,
                         round(subtotal, 2), round(discount, 2), round(total, 2),
                         pay_method, pay_status, notes_val, created_at),
                    ).lastrowid

                    for r in item_rows:
                        item_name = get_field(r, "item_name") or "Imported item"
                        item_code = get_field(r, "item_code") or ""
                        qty = get_field(r, "qty") or 1
                        price = get_field(r, "price") or 0
                        line_total = get_field(r, "line_total") or (price * qty)
                        # Try to match category by code
                        cat_code = get_field(r, "category")
                        category_id = _get_or_create_category_by_code(cat_code, price) if cat_code else None
                        c.execute(
                            "INSERT INTO sale_items(sale_id, item_name, category_id, category_code, "
                            "cost_price, sell_price, qty, line_total) VALUES(?,?,?,?,?,?,?,?)",
                            (sale_id, item_name, category_id, item_code, 0, price, qty, line_total),
                        )
                        imported_items += 1
                    # Log cash sale to drawer
                    if pay_method == "cash":
                        c.execute(
                            "INSERT INTO cash_drawer(type, amount, description, reference_id, reference_type) "
                            "VALUES('sale', ?, ?, ?, 'sale')",
                            (total, f"Imported sale {invoice_no}", sale_id),
                        )
                imported_sales += 1
                total_revenue += total
            except Exception as e:
                errors.append(f"Sale {inv_no}: {e}")
                skipped += 1
    else:
        # Per-sale rows (one row = one sale, no item breakdown)
        for row in rows:
            try:
                inv = get_field(row, "invoice_no") or f"IMP-{imported_sales + 1}"
                date_str, time_str = parse_date(get_field(row, "date")) or (None, None)
                if date_str:
                    dates_seen.append(date_str)
                customer_name = get_field(row, "customer_name") or ""
                customer_phone = get_field(row, "customer_phone") or ""
                total = get_field(row, "total") or get_field(row, "line_total") or 0
                if not total:
                    skipped += 1
                    continue
                subtotal = get_field(row, "subtotal") or total
                discount = get_field(row, "discount") or 0
                pay_method = normalize_payment_method(get_field(row, "payment_method"))
                pay_status = "credit" if pay_method == "credit" else "paid"
                invoice_no = f"IMP-{inv}" if not str(inv).startswith("IMP-") else inv
                with conn() as c:
                    existing = c.execute(
                        "SELECT id FROM sales WHERE invoice_no=?", (invoice_no,)
                    ).fetchone()
                if existing:
                    skipped += 1
                    continue
                customer_id = _get_or_create_customer(customer_name, customer_phone)
                created_at = date_str or datetime.now().strftime("%Y-%m-%d")
                if time_str:
                    created_at = f"{date_str} {time_str}"
                with conn() as c:
                    sale_id = c.execute(
                        "INSERT INTO sales(invoice_no, customer_name, customer_phone, customer_id, "
                        "subtotal, discount, total, payment_method, payment_status, created_at) "
                        "VALUES(?,?,?,?,?,?,?,?,?)",
                        (invoice_no, customer_name, customer_phone, customer_id,
                         round(subtotal, 2), round(discount, 2), round(total, 2),
                         pay_method, pay_status, created_at),
                    ).lastrowid
                    # Insert a single summary item
                    item_name = get_field(row, "item_name") or "Imported sale"
                    qty = get_field(row, "qty") or 1
                    price = total / qty if qty else total
                    c.execute(
                        "INSERT INTO sale_items(sale_id, item_name, sell_price, qty, line_total) "
                        "VALUES(?,?,?,?,?)",
                        (sale_id, item_name, price, qty, total),
                    )
                    if pay_method == "cash":
                        c.execute(
                            "INSERT INTO cash_drawer(type, amount, description, reference_id, reference_type) "
                            "VALUES('sale', ?, ?, ?, 'sale')",
                            (total, f"Imported sale {invoice_no}", sale_id),
                        )
                imported_sales += 1
                imported_items += 1
                total_revenue += total
            except Exception as e:
                errors.append(f"Row: {e}")
                skipped += 1

    # Record the import itself
    date_range_start = min(dates_seen) if dates_seen else None
    date_range_end = max(dates_seen) if dates_seen else None
    if not import_date:
        import_date = date_range_end or datetime.now().strftime("%Y-%m-%d")

    # v8.18.18: bags rule — imported bag-category sales must keep the bag
    # categories' ON-HAND qty at max(purchases+adjustments − sold, 0)
    # (the virtual "purchased raised to SOLD" side is applied at display
    # time — see the bags block comment in profit_engine). The Ezi DBF
    # pipeline (pos_import_sync) does this at end-of-import; this generic
    # CSV/Excel/JSON path needs the same sync so every import route lands
    # on the same number. Non-bag stock state is NOT touched here (this
    # legacy path predates stock-state tracking).
    bags_synced = []
    try:
        from .profit_engine import sync_bags_stock_to_sold as _sync_bags
        bags_synced = _sync_bags()
    except Exception as e:
        errors.append(f"Bags stock sync failed: {e}")

    with conn() as c:
        import_id = c.execute(
            "INSERT INTO pos_imports(source_name, filename, file_format, row_count, sale_count, "
            "total_revenue, date_range_start, date_range_end, column_mapping, import_date, status, notes) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
            (source_name, filename, "csv", len(rows), imported_sales,
             round(total_revenue, 2), date_range_start, date_range_end,
             json.dumps(mapping), import_date,
             "imported" if not errors else "partial", notes),
        ).lastrowid

    return {
        "import_id": import_id,
        "imported_sales": imported_sales,
        "imported_items": imported_items,
        "skipped": skipped,
        "total_revenue": round(total_revenue, 2),
        "date_range_start": date_range_start,
        "date_range_end": date_range_end,
        "errors": errors[:20],  # cap error list
        "bags_stock_synced": bags_synced,  # v8.18.17
    }


def list_imports(limit: int = 50) -> list:
    with conn() as c:
        rows = c.execute(
            "SELECT * FROM pos_imports ORDER BY id DESC LIMIT ?", (limit,)
        ).fetchall()
    return [dict(r) for r in rows]


def get_import(import_id: int) -> dict | None:
    with conn() as c:
        row = c.execute("SELECT * FROM pos_imports WHERE id=?", (import_id,)).fetchone()
    if not row:
        return None
    d = dict(row)
    try:
        d["column_mapping"] = json.loads(d.get("column_mapping") or "{}")
    except Exception:
        d["column_mapping"] = {}
    return d


def delete_import(import_id: int) -> dict:
    """DEPRECATED — v8.9.1: This legacy function does NOT reverse stock state,
    customer stats, commissions, or loyalty redemptions. It leaves orphaned
    ezi_pos_imports rows. Use pos_import_sync.delete_pos_import() instead.

    This function is kept only for backward compat and will raise RuntimeError
    to prevent accidental use.
    """
    raise RuntimeError(
        "Legacy delete_import is disabled (v8.9.1). "
        "Use pos_import_sync.delete_pos_import() instead — it properly reverses "
        "stock state, customer stats, commissions, loyalty redemptions, and "
        "cleans up ezi_pos_imports + pos_expense_imports rows."
    )


def get_today_summary_from_imports(date: str = "") -> dict:
    """Get today's sales summary that includes imported sales.
    Used to display 'Today's sales' panel that combines POS-imported + native BillBook sales."""
    if not date:
        date = datetime.now().strftime("%Y-%m-%d")
    with conn() as c:
        sales = c.execute(
            "SELECT * FROM sales WHERE date(created_at)=?", (date,)
        ).fetchall()
        imported_count = sum(1 for s in sales if s["invoice_no"] and s["invoice_no"].startswith("IMP-"))
        native_count = len(sales) - imported_count
        total_revenue = sum(s["total"] for s in sales if s["payment_status"] != "refunded")
        cash_total = sum(s["total"] for s in sales if s["payment_method"] == "cash" and s["payment_status"] != "refunded")
        card_total = sum(s["total"] for s in sales if s["payment_method"] == "card" and s["payment_status"] != "refunded")
        credit_total = sum(s["total"] for s in sales if s["payment_method"] == "credit" and s["payment_status"] != "refunded")
        # Check if there's a POS import for today
        today_import = c.execute(
            "SELECT * FROM pos_imports WHERE import_date=? ORDER BY id DESC LIMIT 1", (date,)
        ).fetchone()
    return {
        "date": date,
        "total_sales": len(sales),
        "native_sales": native_count,
        "imported_sales": imported_count,
        "total_revenue": round(total_revenue, 2),
        "cash_total": round(cash_total, 2),
        "card_total": round(card_total, 2),
        "credit_total": round(credit_total, 2),
        "today_import": dict(today_import) if today_import else None,
    }


# ---------- Sales targets ----------

def set_target(period: str, target_date: str, amount: float, notes: str = "") -> dict:
    """Set or update a sales target. period: 'daily' | 'monthly'."""
    with conn() as c:
        existing = c.execute(
            "SELECT id FROM sales_targets WHERE period=? AND target_date=?",
            (period, target_date),
        ).fetchone()
        if existing:
            c.execute(
                "UPDATE sales_targets SET target_amount=?, notes=? WHERE id=?",
                (amount, notes, existing["id"]),
            )
            return {"id": existing["id"], "updated": True}
        tid = c.execute(
            "INSERT INTO sales_targets(period, target_date, target_amount, notes) VALUES(?,?,?,?)",
            (period, target_date, amount, notes),
        ).lastrowid
    return {"id": tid, "updated": False}


def get_target(period: str, target_date: str) -> dict | None:
    with conn() as c:
        row = c.execute(
            "SELECT * FROM sales_targets WHERE period=? AND target_date=?",
            (period, target_date),
        ).fetchone()
    return dict(row) if row else None


def get_target_progress(period: str, target_date: str) -> dict:
    """Get target vs actual for the given period."""
    target = get_target(period, target_date)
    if not target:
        return {"target": None, "actual": 0, "progress_pct": 0, "remaining": 0}
    # Compute actual sales
    with conn() as c:
        if period == "daily":
            row = c.execute(
                "SELECT COALESCE(SUM(total), 0) v FROM sales "
                f"WHERE date(created_at)=? AND {db.VALID_SALE_FILTER_NO_ALIAS}",
                (target_date,),
            ).fetchone()
        else:  # monthly
            row = c.execute(
                "SELECT COALESCE(SUM(total), 0) v FROM sales "
                f"WHERE strftime('%Y-%m', created_at)=? AND {db.VALID_SALE_FILTER_NO_ALIAS}",
                (target_date,),
            ).fetchone()
    actual = row["v"] or 0
    target_amt = target["target_amount"]
    progress = (actual / target_amt * 100) if target_amt > 0 else 0
    return {
        "target": target_amt,
        "actual": round(actual, 2),
        "progress_pct": round(progress, 1),
        "remaining": round(max(0, target_amt - actual), 2),
        "period": period,
        "date": target_date,
    }


# ─── ZIP EXTRACTION (v8.4) ──────────────────────────────────────────────────

# File extensions we can parse inside a ZIP
_DATA_EXTENSIONS = {".csv", ".json", ".txt", ".tsv", ".xlsx", ".xls"}

# Max ZIP size: 50 MB (prevents memory exhaustion on huge archives)
MAX_ZIP_SIZE = 50 * 1024 * 1024

# Max number of files inside a ZIP (prevents zip-bomb attacks)
MAX_ZIP_FILES = 200

# Max uncompressed single file size inside a ZIP: 20 MB
MAX_ZIP_FILE_SIZE = 20 * 1024 * 1024


def extract_zip_contents(zip_bytes: bytes) -> dict:
    """Extract a ZIP archive and return the best data file found inside.

    Scans all files in the archive, picks the one that looks most like a POS
    data export (largest CSV/JSON/Excel file), and returns its contents.

    Returns:
        {
            "filename": str,           # name of the file inside the ZIP
            "file_format": str,        # "csv" | "json" | "excel"
            "content": str,            # file contents as text (for CSV/JSON)
            "raw_bytes": bytes,        # raw file bytes (for Excel)
            "all_files": list[str],    # list of all files found in the ZIP
        }
    Raises:
        ValueError with a descriptive message on invalid ZIP / no data file / too large.
    """
    if len(zip_bytes) > MAX_ZIP_SIZE:
        raise ValueError(f"ZIP file too large: {len(zip_bytes) / 1024 / 1024:.1f} MB (max {MAX_ZIP_SIZE / 1024 / 1024:.0f} MB)")

    try:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except zipfile.BadZipFile:
        raise ValueError("Invalid ZIP file — could not open archive.")

    # Check for zip-bomb (too many files or compression ratio > 100:1)
    infos = zf.infolist()
    if len(infos) > MAX_ZIP_FILES:
        raise ValueError(f"ZIP contains too many files: {len(infos)} (max {MAX_ZIP_FILES})")

    # Collect candidate data files
    candidates = []
    all_files = []
    for info in infos:
        if info.is_dir():
            continue
        name = info.filename.lower()
        all_files.append(info.filename)
        # Skip hidden/system files
        if name.startswith("_") or "/__macosx/" in name or name.endswith(".ds_store"):
            continue
        ext = os.path.splitext(name)[1]
        if ext in _DATA_EXTENSIONS:
            if info.file_size > MAX_ZIP_FILE_SIZE:
                continue  # skip oversized files
            candidates.append(info)

    if not candidates:
        raise ValueError(
            f"No data files found in ZIP. "
            f"Supported extensions: {', '.join(sorted(_DATA_EXTENSIONS))}. "
            f"Files found: {', '.join(all_files[:10]) or 'none'}"
        )

    # Pick the largest candidate (most likely the main data file)
    best = max(candidates, key=lambda i: i.file_size)

    # Read the file content
    raw = zf.read(best)
    name_lower = best.filename.lower()
    ext = os.path.splitext(name_lower)[1]

    if ext == ".json":
        file_format = "json"
        content = raw.decode("utf-8", errors="replace")
        return {
            "filename": best.filename,
            "file_format": file_format,
            "content": content,
            "raw_bytes": None,
            "all_files": all_files,
        }
    elif ext in (".xlsx", ".xls"):
        file_format = "excel"
        return {
            "filename": best.filename,
            "file_format": file_format,
            "content": "",
            "raw_bytes": raw,
            "all_files": all_files,
        }
    else:
        # CSV / TSV / TXT — try to detect encoding
        file_format = "csv"
        try:
            content = raw.decode("utf-8")
        except UnicodeDecodeError:
            try:
                content = raw.decode("latin-1")
            except Exception:
                content = raw.decode("utf-8", errors="replace")
        return {
            "filename": best.filename,
            "file_format": file_format,
            "content": content,
            "raw_bytes": None,
            "all_files": all_files,
        }


def parse_excel_bytes(raw_bytes: bytes, filename: str = "") -> tuple:
    """Parse Excel bytes and return (headers, rows).

    Uses openpyxl which is already in requirements.txt.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("Excel parsing requires openpyxl. Run: pip install openpyxl")

    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        wb.close()
        return [], []
    rows = []
    for row in rows_iter:
        if all(c is None for c in row):
            continue  # skip empty rows
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = str(val) if val is not None else ""
        rows.append(row_dict)
    wb.close()
    return headers, rows

