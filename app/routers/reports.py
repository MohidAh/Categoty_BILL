"""Auto-generated router module — extracted from main.py Phase 1."""
import os, json, time, re, io, csv, secrets, hashlib, traceback
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import quote

from fastapi import APIRouter, Request, HTTPException, UploadFile, File, Form, BackgroundTasks
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse, HTMLResponse, RedirectResponse
from pydantic import BaseModel
from typing import Any, Optional

from .. import db
from .. import shop as shop_mod
from .. import insights
from .. import trends as trends_mod
from .. import extract
from .. import reports
from .. import pos_extra
from .. import pos_import
from .. import crypto as crypto_mod
from .. import jobs as jobs_mod
from ..config import BACKUPS, BASE, PAGE_SIZE, PAGES, UPLOADS
from ..export import export_bills, export_insights
from ..ingest import render_pages, save_upload
from ..validate import detect_duplicate, pieces, validate
from ..security import (
    hash_password, verify_password, ensure_password,
    is_logged_in, get_session, get_session_role,
    create_session, delete_session,
    check_login_throttle, record_failed_login,
    SESSION_DAYS,
)

router = APIRouter()

# Backward-compat aliases
_hash_password = hash_password
_verify_password = verify_password
_ensure_password = ensure_password
_is_logged_in = is_logged_in
_get_session = get_session
_get_session_role = get_session_role
_create_session = create_session
_delete_session = delete_session
_check_login_throttle = check_login_throttle
_record_failed_login = record_failed_login

MAX_UPLOAD_BYTES = 200 * 1024 * 1024
MAX_FILES = 100


@router.get("/api/reports/monthly")
def r_monthly(start: str, end: str) -> Any:
    return reports.monthly_summary(start, end)




@router.get("/api/reports/billwise")
def r_billwise(start: str, end: str, status: str = "all",
                include_items: bool = False) -> Any:
    """Bill-by-bill report.

    v8.5.2: status filter — 'all' (default), 'confirmed', or 'review'.
    Default is 'all' so newly-uploaded bills appear immediately, even before
    the user has reviewed + confirmed them.

    v8.7: by default returns bill headers + precomputed aggregates (NO items
    array) for the lazy-load master list. Pass include_items=true for the
    legacy behavior (all items in one payload — used by Excel export).
    For single-bill detail, use GET /api/bills/{bill_id}.
    """
    return reports.billwise_report(start, end, status=status,
                                    include_items=include_items)


@router.get("/api/reports/billwise/export")
def r_billwise_export(start: str = "", end: str = "", bill_ids: str = "", status: str = "all",
                      format: str = "xlsx") -> Any:
    """v8.5.3: Export selected bills as Excel (.xlsx) — per-category sheets.

    Workbook structure (matches user's reference Excel):
      - Sheet "250":  all line items where category sell_price=250
      - Sheet "500":  all line items where category sell_price=500
      - Sheet "750":  all line items where category sell_price=750
      - Sheet "1000": all line items where category sell_price=1000
      - Sheet "Other": items with any other sell_price (e.g. bags, custom)
      - Sheet "Summary": per-category totals (count, total qty, total amount)

    Each per-category sheet has 4 simple columns:
      Sr. no. | Price | Quantity | total Amount

    If bill_ids is provided (comma-separated), only those bills are exported.
    Otherwise all bills in the date range + status filter are exported.

    v8.16.4: format=pdf now delegates to the universal PDF generator
    (the per-category Excel workbook only makes sense for Excel).

    v8.16.7: `start` and `end` are now optional (default to last 30 days)
    — fixes 422 error when PDF button is clicked without date params.
    """
    # v8.16.7: Default to last 30 days if not provided
    if not start or not end:
        from datetime import datetime as _dt_def, timedelta as _td
        today = _dt_def.now().date()
        if not start:
            start = (today - _td(days=30)).strftime("%Y-%m-%d")
        if not end:
            end = today.strftime("%Y-%m-%d")

    # v8.16.4: For PDF, fall back to the universal generator
    if format.lower() == "pdf":
        data = reports.billwise_report(start, end, status)
        from datetime import datetime as _dt3
        ts = _dt3.now().strftime("%Y%m%d_%H%M%S")
        filename = f"billwise_{ts}"
        return _generate_pdf("billwise", data, filename)
    from ..export import _header_style
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    from fastapi.responses import StreamingResponse

    # Get the billwise data
    d = reports.billwise_report(start, end, status=status, include_items=True)
    bills = d.get("bills", [])

    # Filter to selected bill IDs if provided
    if bill_ids:
        selected_ids = set()
        for bid in bill_ids.split(","):
            try:
                selected_ids.add(int(bid.strip()))
            except ValueError:
                pass
        bills = [b for b in bills if b["bill_id"] in selected_ids]

    # Group items by category sell_price
    # Use a dict: {sell_price: [{sr_no, price, qty, line_total, bill_id, bill_date, supplier}, ...]}
    items_by_price = {}
    for b in bills:
        for it in b["items"]:
            sp = it.get("sell_price") or 0
            # Round to nearest int for the sheet name (250.0 → "250")
            try:
                sp_key = int(sp) if sp == int(sp) else sp
            except (TypeError, ValueError):
                sp_key = sp
            items_by_price.setdefault(sp_key, []).append({
                "sr_no": it.get("sr_no", 0),
                "price": it.get("price", 0),
                "qty": it.get("qty", 0),
                "pieces": it.get("pieces", 0),
                "line_total": it.get("line_total", 0),
                "bill_id": b["bill_id"],
                "bill_date": b.get("bill_date", ""),
                "supplier": b.get("supplier_name", ""),
                "raw": it.get("raw", ""),
                "item_code": it.get("item_code", ""),
            })

    # Standard price tiers in order
    standard_prices = [250, 500, 750, 1000]
    # Other prices (e.g. bags at 10/20/30/50/60)
    other_prices = sorted([p for p in items_by_price.keys() if p not in standard_prices])

    wb = Workbook()
    # Remove the default sheet — we'll create our own
    wb.remove(wb.active)

    # Subtle border for cells
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB'),
    )
    # Header style: bold white text on dark background
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F2A44")
    header_align = Alignment(horizontal="center", vertical="center")
    # Number style: black text, right-aligned
    num_align = Alignment(horizontal="right")

    def write_category_sheet(sheet_name: str, items: list):
        """Write one per-category sheet with columns: Sr. no. | Price | Quantity | total Amount"""
        ws = wb.create_sheet(sheet_name)
        headers = ["Sr. no.", "Price", "Quantity", "total Amount"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # If no items, add a placeholder row so the sheet isn't empty
        if not items:
            ws.cell(row=2, column=1, value="No items in this price tier").font = Font(italic=True, color="64748B")
            ws.column_dimensions['A'].width = 35
            return

        for idx, it in enumerate(items, 2):
            # Use pieces (handles dozen conversion) as the quantity column
            qty = it["pieces"] if it["pieces"] else it["qty"]
            line_total = it["line_total"] if it["line_total"] else (it["price"] * qty)
            ws.cell(row=idx, column=1, value=idx - 1).border = thin_border  # Sr. no.
            price_cell = ws.cell(row=idx, column=2, value=it["price"])
            price_cell.number_format = '0.00'
            price_cell.border = thin_border
            price_cell.alignment = num_align
            qty_cell = ws.cell(row=idx, column=3, value=qty)
            qty_cell.number_format = '0'
            qty_cell.border = thin_border
            qty_cell.alignment = num_align
            total_cell = ws.cell(row=idx, column=4, value=round(line_total, 2))
            total_cell.number_format = '#,##0.00'
            total_cell.border = thin_border
            total_cell.alignment = num_align

        # Totals row at the bottom
        total_row = len(items) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = thin_border
        total_qty = sum((it["pieces"] or it["qty"]) for it in items)
        total_amt = sum((it["line_total"] or (it["price"] * (it["pieces"] or it["qty"]))) for it in items)
        ws.cell(row=total_row, column=3, value=total_qty).font = Font(bold=True)
        ws.cell(row=total_row, column=3).alignment = num_align
        ws.cell(row=total_row, column=3).border = thin_border
        ws.cell(row=total_row, column=4, value=round(total_amt, 2)).font = Font(bold=True)
        ws.cell(row=total_row, column=4).number_format = '#,##0.00'
        ws.cell(row=total_row, column=4).alignment = num_align
        ws.cell(row=total_row, column=4).border = thin_border

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 16
        ws.freeze_panes = "A2"

    # Create per-category sheets in order: 250, 500, 750, 1000, then any others
    for price in standard_prices:
        items = items_by_price.get(price, [])
        if items:  # only create the sheet if it has items
            write_category_sheet(str(price), items)
    for price in other_prices:
        items = items_by_price.get(price, [])
        if items:
            write_category_sheet(str(price), items)

    # Summary sheet — last (matches the reference Excel layout)
    ws_summary = wb.create_sheet("Summary")
    summary_headers = ["Category", "Items", "Total Qty", "Total Amount", "% of Total"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    all_prices_in_order = standard_prices + other_prices
    grand_total_amount = sum(
        sum((it["line_total"] or (it["price"] * (it["pieces"] or it["qty"]))) for it in items_by_price.get(p, []))
        for p in all_prices_in_order
    )
    row_idx = 2
    for price in all_prices_in_order:
        items = items_by_price.get(price, [])
        if not items:
            continue
        total_qty = sum((it["pieces"] or it["qty"]) for it in items)
        total_amt = sum((it["line_total"] or (it["price"] * (it["pieces"] or it["qty"]))) for it in items)
        pct = (total_amt / grand_total_amount * 100) if grand_total_amount > 0 else 0
        ws_summary.cell(row=row_idx, column=1, value=str(price)).border = thin_border
        ws_summary.cell(row=row_idx, column=2, value=len(items)).border = thin_border
        ws_summary.cell(row=row_idx, column=2).alignment = num_align
        ws_summary.cell(row=row_idx, column=3, value=total_qty).border = thin_border
        ws_summary.cell(row=row_idx, column=3).alignment = num_align
        amt_cell = ws_summary.cell(row=row_idx, column=4, value=round(total_amt, 2))
        amt_cell.number_format = '#,##0.00'
        amt_cell.border = thin_border
        amt_cell.alignment = num_align
        pct_cell = ws_summary.cell(row=row_idx, column=5, value=f"{pct:.1f}%")
        pct_cell.border = thin_border
        pct_cell.alignment = num_align
        row_idx += 1

    # Grand total row
    ws_summary.cell(row=row_idx, column=1, value="GRAND TOTAL").font = Font(bold=True)
    ws_summary.cell(row=row_idx, column=1).border = thin_border
    ws_summary.cell(row=row_idx, column=2, value=sum(len(items_by_price.get(p, [])) for p in all_prices_in_order)).font = Font(bold=True)
    ws_summary.cell(row=row_idx, column=2).alignment = num_align
    ws_summary.cell(row=row_idx, column=2).border = thin_border
    ws_summary.cell(row=row_idx, column=3, value=sum(sum((it["pieces"] or it["qty"]) for it in items_by_price.get(p, [])) for p in all_prices_in_order)).font = Font(bold=True)
    ws_summary.cell(row=row_idx, column=3).alignment = num_align
    ws_summary.cell(row=row_idx, column=3).border = thin_border
    gt_cell = ws_summary.cell(row=row_idx, column=4, value=round(grand_total_amount, 2))
    gt_cell.font = Font(bold=True)
    gt_cell.number_format = '#,##0.00'
    gt_cell.alignment = num_align
    gt_cell.border = thin_border
    ws_summary.cell(row=row_idx, column=5, value="100.0%").font = Font(bold=True)
    ws_summary.cell(row=row_idx, column=5).border = thin_border

    ws_summary.column_dimensions['A'].width = 12
    ws_summary.column_dimensions['B'].width = 10
    ws_summary.column_dimensions['C'].width = 12
    ws_summary.column_dimensions['D'].width = 18
    ws_summary.column_dimensions['E'].width = 14
    ws_summary.freeze_panes = "A2"

    # If no sheets were created (no items at all), add a placeholder
    if len(wb.worksheets) == 0:
        ws = wb.create_sheet("No Data")
        ws.cell(row=1, column=1, value="No bills in the selected date range with the chosen filters.")
        ws.cell(row=2, column=1, value="Try widening the date range or changing the filters.")

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"billwise_{start}_to_{end}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )




@router.get("/api/reports/category")
def r_category(start: str = "", end: str = "") -> Any:
    """Category-wise report: total products, total cost, avg amount, profit margin per category."""
    return reports.category_report(start or None, end or None)




@router.get("/api/reports/profit")
def r_profit(start: str, end: str) -> Any:
    return reports.profit_estimate(start, end)




@router.get("/api/reports/suppliers")
def r_suppliers() -> Any:
    return reports.supplier_ranking()


# ------------------------------------------------------------------
# Insights
# ------------------------------------------------------------------



@router.get("/api/reports/monthly-close")
def monthly_close_route(year: int, month: int) -> Any:
    """Snapshot all data for a specific month for accounting closure.
    v8.2: also triggers a month-end audit run."""
    return insights.monthly_close_with_audit(year, month)




@router.get("/api/reports/monthly-close.pdf")
def monthly_close_pdf(year: int, month: int) -> Any:
    """Generate a PDF report for monthly closing."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                     TableStyle)
    from reportlab.lib.units import mm
    import io as _io

    data = insights.monthly_close(year, month)
    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm,
                            leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'],
                                  fontSize=18, spaceAfter=10)
    heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'],
                                    fontSize=13, spaceAfter=6, spaceBefore=12)
    normal_style = ParagraphStyle('CustomNormal', parent=styles['Normal'],
                                   fontSize=10, spaceAfter=4)

    elements = []
    elements.append(Paragraph(f"BillBook — Monthly Report", title_style))
    elements.append(Paragraph(f"<b>Period:</b> {data['month']}", normal_style))
    elements.append(Paragraph(f"<b>Total Bills:</b> {data['total_bills']}", normal_style))
    elements.append(Paragraph(f"<b>Total Spent:</b> Rs {data['total_spent']:,.0f}", normal_style))
    elements.append(Paragraph(f"<b>Paid:</b> Rs {data['total_paid']:,.0f}  |  "
                              f"<b>Credit:</b> Rs {data['total_credit']:,.0f}", normal_style))
    elements.append(Paragraph(f"<b>Suppliers:</b> {data['supplier_count']}", normal_style))
    elements.append(Spacer(1, 10*mm))

    # v8.18.14 — Sales & Income Summary. The dedicated PDF used to show
    # ONLY the buy side (bills); the sell side existed in the JSON but never
    # rendered here. Extra (non-POS) sales get their own clearly-labeled
    # line so they are differentiable from POS sales revenue.
    elements.append(Paragraph("Sales &amp; Income Summary", heading_style))
    _es = data.get("extra_sales_income") or 0
    _es_n = data.get("extra_sales_count") or 0
    summary_rows = [
        ["Item", "Amount"],
        ["POS Sales Revenue (net of discounts)", f"Rs {data['total_revenue']:,.0f}"],
        ["POS Sales Invoices", str(data['sales_count'])],
        ["Extra Sales — non-POS (cartons, raddi…)",
         f"Rs {_es:,.0f}  ({_es_n} entries)"],
        ["Cost of Goods Sold (POS)", f"Rs {data['cost_of_goods']:,.0f}"],
        ["Gross Profit (POS revenue − COGS)", f"Rs {data['gross_profit']:,.0f}"],
        ["Operating Expenses", f"Rs {data['operating_expenses']:,.0f}"],
        ["Net Profit (gross + extra sales − op. expenses)",
         f"Rs {data['net_profit']:,.0f}"],
        ["Owner Draws (equity, not expense)", f"Rs {data.get('owner_draws') or 0:,.0f}"],
    ]
    t_sum = Table(summary_rows, colWidths=[120*mm, 55*mm])
    t_sum.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        # highlight the extra-sales row so it stands out from POS lines
        ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#ecfdf5')),
    ]))
    elements.append(t_sum)
    elements.append(Spacer(1, 8*mm))

    # v8.18.14 — Extra Sales line items table (only when entries exist)
    if data.get("extra_sales"):
        elements.append(Paragraph("Extra Sales Entries (non-POS)", heading_style))
        es_rows = [["Date", "Item", "Description", "Qty", "Rate", "Total"]]
        for e in data["extra_sales"]:
            es_rows.append([
                (e["sale_date"] or "")[:10],
                e["item_name"] or "—",
                (e["description"] or "—")[:40],
                f"{e['quantity']:g}",
                f"Rs {e['unit_price']:,.0f}",
                f"Rs {e['total']:,.0f}",
            ])
        es_rows.append(["", "", "", "", "TOTAL", f"Rs {_es:,.0f}"])
        t_es = Table(es_rows, colWidths=[22*mm, 38*mm, 55*mm, 15*mm, 22*mm, 23*mm])
        t_es.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(t_es)
        elements.append(Spacer(1, 8*mm))

    # Bills table
    elements.append(Paragraph("Bills", heading_style))
    bill_rows = [["ID", "Date", "Supplier", "Total", "Payment"]]
    for b in data["bills"]:
        bill_rows.append([
            str(b["id"]),
            b["bill_date"][:10] if b["bill_date"] else "—",
            b["supplier_name"] or "—",
            f"Rs {(b['written_total'] or b['computed_total'] or 0):,.0f}",
            b["payment_status"],
        ])
    t = Table(bill_rows, colWidths=[15*mm, 25*mm, 60*mm, 30*mm, 25*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 8*mm))

    # Category breakdown
    if data["by_category"]:
        elements.append(Paragraph("Category Breakdown", heading_style))
        cat_rows = [["Category", "Items", "Pieces", "Cost", "Revenue"]]
        for cat, vals in sorted(data["by_category"].items()):
            cat_rows.append([
                cat, str(vals["items"]), str(vals["pieces"]),
                f"Rs {vals['cost']:,.0f}", f"Rs {vals['revenue']:,.0f}",
            ])
        t2 = Table(cat_rows, colWidths=[40*mm, 20*mm, 25*mm, 35*mm, 35*mm])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        elements.append(t2)

    doc.build(elements)
    buf.seek(0)
    filename = f"billbook_monthly_{data['month']}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )




@router.get("/api/export/bills.xlsx")
def r_export_bills() -> Any:
    data = export_bills()
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=billbook_bills.xlsx"},
    )




@router.get("/api/export/insights.xlsx")
def r_export_insights() -> Any:
    data = export_insights()
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=billbook_insights.xlsx"},
    )




@router.get("/api/export.csv")
def export_csv() -> Any:
    """Legacy flat CSV for backward compatibility."""
    out = io.StringIO()
    out.write("\ufeff")  # BOM for Excel
    w = csv.writer(out)
    w.writerow(["bill_id", "supplier", "phone", "date", "raw", "price", "qty",
                "unit", "pieces", "line_total", "bill_total", "status", "payment"])
    with db.conn() as c:
        for b in c.execute("SELECT * FROM bills WHERE deleted_at IS NULL ORDER BY COALESCE(bill_date, date(created_at)) DESC, id DESC"):
            for i in c.execute(
                "SELECT * FROM bill_items WHERE bill_id=? ORDER BY id", (b["id"],)
            ):
                w.writerow([
                    b["id"], b["supplier_name"], b["phone"], b["bill_date"], i["raw"],
                    i["price"], i["qty"], i["unit"], pieces(i["qty"], i["unit"]),
                    i["line_total"], b["written_total"], b["status"], b["payment_status"],
                ])
    out.seek(0)
    return StreamingResponse(
        iter([out.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=bills.csv"},
    )


# ------------------------------------------------------------------
# AI Providers (Settings)
# ------------------------------------------------------------------



@router.get("/api/reports/pnl")
def pnl_report(month: str = "") -> Any:
    return shop_mod.get_pnl(month)


# v4.0 Phase 3 — Actual Earnings (hero dashboard)
@router.get("/api/reports/actual-earnings")
def actual_earnings_report(month: str = "") -> Any:
    """The single 'truth' dashboard. Returns total_sales, cogs, gross_profit,
    operating_expenses, actual_earnings, net_margin, purchases (shown separately),
    expenses_by_category, cash_reality, and comparison vs last month.
    """
    return shop_mod.get_actual_earnings(month)


# ------------------------------------------------------------------
# Held Orders (park & recall)
# ------------------------------------------------------------------



@router.get("/api/reports/cash-flow")
def cash_flow_report(month: str = "") -> Any:
    return pos_extra.get_cash_flow(month)




@router.get("/api/reports/balance-sheet")
def balance_sheet_report(as_of: str = "") -> Any:
    return pos_extra.get_balance_sheet(as_of)


# ==================================================================
# Barcode / QR codes
# ==================================================================



@router.get("/api/reports/top-items")
def top_items_report(start: str = "", end: str = "", limit: int = 20) -> Any:
    """Top-selling items by revenue in a date range."""
    if not start:
        start = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end:
        end = datetime.now().strftime("%Y-%m-%d")
    with db.conn() as c:
        rows = c.execute(
            "SELECT si.item_name, si.category_code, "
            "SUM(si.qty) AS total_qty, "
            "SUM(si.line_total) AS total_revenue, "
            "COUNT(DISTINCT s.id) AS sale_count, "
            "AVG(si.sell_price) AS avg_price "
            "FROM sale_items si JOIN sales s ON si.sale_id = s.id "
            "WHERE date(s.created_at) BETWEEN ? AND ? "
            f"AND {db.VALID_SALE_FILTER} "
            "GROUP BY si.item_name, si.category_code "
            "ORDER BY total_revenue DESC LIMIT ?",
            (start, end, limit),
        ).fetchall()
    return {
        "start": start, "end": end,
        "items": [dict(r) for r in rows],
    }


# ==================================================================
# Peak hours report (sales by hour heatmap data)
# ==================================================================



@router.get("/api/reports/peak-hours")
def peak_hours_report(start: str = "", end: str = "") -> Any:
    """Sales count + revenue by hour of day, for heatmap visualization."""
    if not start:
        start = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end:
        end = datetime.now().strftime("%Y-%m-%d")
    with db.conn() as c:
        rows = c.execute(
            "SELECT CAST(substr(created_at, 12, 2) AS INTEGER) AS hour, "
            "COUNT(*) AS sale_count, "
            "SUM(total) AS revenue "
            "FROM sales "
            "WHERE date(created_at) BETWEEN ? AND ? "
            f"AND {db.VALID_SALE_FILTER_NO_ALIAS} "
            "GROUP BY hour ORDER BY hour",
            (start, end),
        ).fetchall()
    # Build full 24-hour array
    by_hour = {r["hour"]: {"sale_count": r["sale_count"], "revenue": round(r["revenue"] or 0, 2)}
               for r in rows}
    full = [{"hour": h, **by_hour.get(h, {"sale_count": 0, "revenue": 0})} for h in range(24)]
    peak_hour = max(rows, key=lambda r: r["sale_count"]) if rows else None
    return {
        "start": start, "end": end,
        "by_hour": full,
        "peak_hour": peak_hour["hour"] if peak_hour else None,
        "peak_count": peak_hour["sale_count"] if peak_hour else 0,
        "total_sales": sum(r["sale_count"] for r in rows),
        "total_revenue": round(sum(r["revenue"] or 0 for r in rows), 2),
    }


# ==================================================================
# Returns & Exchange
# ==================================================================



# ═══════════════════════════════════════════════════
# FBR Digital Invoice Export (Phase 2)
# ═══════════════════════════════════════════════════

@router.get("/api/export/fbr")
def fbr_export(start: str = "", end: str = "") -> Any:
    """Export sales in FBR Digital Invoice format (JSON + CSV).
    Returns JSON array of invoices with: invoice_no, datetime, STRN, NTN, items, tax, total, payment.
    """
    from ..db import get_setting
    if not start:
        start = datetime.now().strftime("%Y-%m-01")
    if not end:
        end = datetime.now().strftime("%Y-%m-%d")

    shop_name = get_setting("shop_name", "BillBook Store")
    ntn = get_setting("shop_ntn", "")
    strn = get_setting("shop_strn", "")

    with db.conn() as c:
        sales = c.execute(
            "SELECT * FROM sales WHERE date(created_at) >= ? AND date(created_at) <= ? "
            f"AND {db.VALID_SALE_FILTER_NO_ALIAS} ORDER BY created_at",
            (start, end),
        ).fetchall()
        invoices = []
        for s in sales:
            items = c.execute(
                "SELECT * FROM sale_items WHERE sale_id=?", (s["id"],)
            ).fetchall()
            invoices.append({
                "invoice_no": s["invoice_no"],
                "datetime": s["created_at"],
                "shop_name": shop_name,
                "ntn": ntn,
                "strn": strn,
                "customer_name": s["customer_name"] or "Walk-in",
                "customer_phone": s["customer_phone"] or "",
                "items": [{
                    "name": it["item_name"],
                    "category_code": it["category_code"],
                    "qty": it["qty"],
                    "price": it["sell_price"],
                    "line_total": it["line_total"],
                } for it in items],
                "subtotal": s["subtotal"],
                "discount": s["discount"],
                "tax_rate": s["tax_rate"],
                "tax_amount": s["tax_amount"],
                "total": s["total"],
                "payment_method": s["payment_method"],
                "payment_status": s["payment_status"],
                "raast_reference": s["raast_reference"] if "raast_reference" in s.keys() else None,
            })

    # Return as JSON (client can also request CSV via ?format=csv)
    return {"invoices": invoices, "count": len(invoices), "period": f"{start} to {end}", "shop_ntn": ntn, "shop_strn": strn}


@router.get("/api/export/fbr.csv")
def fbr_export_csv(start: str = "", end: str = "") -> Any:
    """Export sales as FBR-compatible CSV."""
    from ..db import get_setting
    if not start:
        start = datetime.now().strftime("%Y-%m-01")
    if not end:
        end = datetime.now().strftime("%Y-%m-%d")

    ntn = get_setting("shop_ntn", "")
    strn = get_setting("shop_strn", "")

    with db.conn() as c:
        sales = c.execute(
            "SELECT * FROM sales WHERE date(created_at) >= ? AND date(created_at) <= ? "
            f"AND {db.VALID_SALE_FILTER_NO_ALIAS} ORDER BY created_at",
            (start, end),
        ).fetchall()
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["invoice_no", "datetime", "ntn", "strn", "customer", "phone",
                         "subtotal", "discount", "tax_rate", "tax_amount", "total",
                         "payment_method", "payment_status", "raast_reference"])
        for s in sales:
            writer.writerow([
                s["invoice_no"], s["created_at"], ntn, strn,
                s["customer_name"] or "Walk-in", s["customer_phone"] or "",
                s["subtotal"], s["discount"], s["tax_rate"], s["tax_amount"], s["total"],
                s["payment_method"], s["payment_status"],
                s["raast_reference"] if "raast_reference" in s.keys() else "",
            ])
        content = output.getvalue()

    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=fbr-export-{start}-to-{end}.csv"},
    )


# ═══════════════════════════════════════════════════
# Phase 7: Reports Expansion (12 new)
# ═══════════════════════════════════════════════════

@router.get("/api/reports/ar-aging")
def report_ar_aging() -> Any:
    """AR Aging report — same as /api/customers/ar-aging but in reports namespace."""
    from .customers import ar_aging
    return ar_aging()

@router.get("/api/reports/ap-aging")
def report_ap_aging() -> Any:
    """AP Aging report."""
    from .suppliers import ap_aging
    return ap_aging()

@router.get("/api/reports/inventory-turnover")
def report_inventory_turnover() -> Any:
    """Inventory Turnover report."""
    from .inventory import inventory_kpis
    return inventory_kpis()

@router.get("/api/reports/gmroi")
def report_gmroi() -> Any:
    """GMROI report."""
    from .inventory import inventory_kpis
    kpis = inventory_kpis()
    return {"gmroi": kpis["gmroi"], "cogs_30d": kpis["cogs_30d"], "avg_inventory_value": kpis["avg_inventory_value"]}

@router.get("/api/reports/sell-through")
def report_sell_through() -> Any:
    """Sell-Through % report."""
    from .inventory import inventory_kpis
    kpis = inventory_kpis()
    return {"sell_through_pct": kpis["sell_through_pct"], "received_qty_30d": kpis["received_qty_30d"], "sold_qty_30d": kpis["sold_qty_30d"]}

@router.get("/api/reports/shrinkage")
def report_shrinkage(start: str = "", end: str = "") -> Any:
    """Shrinkage report."""
    from .inventory import shrinkage_report
    return shrinkage_report(start, end)

@router.get("/api/reports/sales-by-customer")
def report_sales_by_customer(start: str = "", end: str = "") -> Any:
    """Top 10 customers by revenue with concentration %."""
    if not start: start = datetime.now().strftime("%Y-%m-01")
    if not end: end = datetime.now().strftime("%Y-%m-%d")
    with db.conn() as c:
        rows = c.execute(
            "SELECT s.customer_id, COALESCE(cust.name, s.customer_name, 'Walk-in') AS name, "
            "COUNT(*) AS sale_count, SUM(s.total) AS revenue "
            "FROM sales s LEFT JOIN customers cust ON s.customer_id = cust.id "
            f"WHERE date(s.created_at) >= ? AND date(s.created_at) <= ? AND {db.VALID_SALE_FILTER} "
            "AND (cust.id IS NULL OR cust.deleted_at IS NULL) "
            "GROUP BY s.customer_id ORDER BY revenue DESC LIMIT 10",
            (start, end),
        ).fetchall()
        total_rev = c.execute(
            f"SELECT COALESCE(SUM(total), 0) AS v FROM sales WHERE date(created_at) >= ? AND date(created_at) <= ? AND {db.VALID_SALE_FILTER_NO_ALIAS}",
            (start, end),
        ).fetchone()["v"]
    customers = [dict(r) for r in rows]
    for c in customers:
        c["concentration_pct"] = round((c["revenue"] / total_rev * 100) if total_rev > 0 else 0, 1)
    return {"customers": customers, "total_revenue": total_rev, "period": f"{start} to {end}"}

@router.get("/api/reports/sales-by-employee")
def report_sales_by_employee(start: str = "", end: str = "") -> Any:
    """Sales by employee/shift."""
    if not start: start = datetime.now().strftime("%Y-%m-01")
    if not end: end = datetime.now().strftime("%Y-%m-%d")
    with db.conn() as c:
        rows = c.execute(
            "SELECT e.id, e.name, e.role, COUNT(s.id) AS tickets, COALESCE(SUM(s.total), 0) AS revenue "
            "FROM employees e LEFT JOIN sales s ON s.employee_id = e.id "
            "AND date(s.created_at) >= ? AND date(s.created_at) <= ? AND " + db.VALID_SALE_FILTER + " "
            "WHERE e.active = 1 GROUP BY e.id ORDER BY revenue DESC",
            (start, end),
        ).fetchall()
    return {"employees": [dict(r) for r in rows], "period": f"{start} to {end}"}

@router.get("/api/reports/atv-basket")
def report_atv_basket(start: str = "", end: str = "") -> Any:
    """Average Transaction Value & Basket Size."""
    if not start: start = datetime.now().strftime("%Y-%m-01")
    if not end: end = datetime.now().strftime("%Y-%m-%d")
    with db.conn() as c:
        row = c.execute(
            "SELECT COUNT(*) AS n, COALESCE(SUM(total), 0) AS rev, COALESCE(AVG(total), 0) AS atv "
            f"FROM sales WHERE date(created_at) >= ? AND date(created_at) <= ? AND {db.VALID_SALE_FILTER_NO_ALIAS}",
            (start, end),
        ).fetchone()
        items_row = c.execute(
            "SELECT COUNT(*) AS item_count, COALESCE(SUM(si.qty), 0) AS total_qty "
            "FROM sale_items si JOIN sales s ON si.sale_id = s.id "
            f"WHERE date(s.created_at) >= ? AND date(s.created_at) <= ? AND {db.VALID_SALE_FILTER}",
            (start, end),
        ).fetchone()
    atv = row["atv"]
    basket_size = (items_row["total_qty"] / row["n"]) if row["n"] > 0 else 0
    return {"atv": round(atv, 2), "basket_size": round(basket_size, 1), "total_sales": row["n"], "total_revenue": row["rev"], "total_items": items_row["total_qty"]}

@router.get("/api/reports/retention")
def report_retention() -> Any:
    """Customer retention: repeat rate 30/60/90 days."""
    with db.conn() as c:
        total_customers = c.execute("SELECT COUNT(*) AS n FROM customers WHERE total_spent > 0 AND deleted_at IS NULL").fetchone()["n"]
        repeat_30 = c.execute(
            "SELECT COUNT(DISTINCT customer_id) AS n FROM sales WHERE customer_id IS NOT NULL "
            "AND date(created_at) >= date('now','-30 days') GROUP BY customer_id HAVING COUNT(*) > 1"
        ).fetchall()
        repeat_60 = c.execute(
            "SELECT COUNT(DISTINCT customer_id) AS n FROM sales WHERE customer_id IS NOT NULL "
            "AND date(created_at) >= date('now','-60 days') GROUP BY customer_id HAVING COUNT(*) > 1"
        ).fetchall()
        repeat_90 = c.execute(
            "SELECT COUNT(DISTINCT customer_id) AS n FROM sales WHERE customer_id IS NOT NULL "
            "AND date(created_at) >= date('now','-90 days') GROUP BY customer_id HAVING COUNT(*) > 1"
        ).fetchall()
    return {
        "total_active_customers": total_customers,
        "repeat_rate_30d": round(len(repeat_30) / max(1, total_customers) * 100, 1),
        "repeat_rate_60d": round(len(repeat_60) / max(1, total_customers) * 100, 1),
        "repeat_rate_90d": round(len(repeat_90) / max(1, total_customers) * 100, 1),
    }

@router.get("/api/reports/supplier-performance")
def report_supplier_performance(start: str = "", end: str = "") -> Any:
    """Supplier performance summary."""
    if not start: start = datetime.now().strftime("%Y-%m-01")
    if not end: end = datetime.now().strftime("%Y-%m-%d")
    with db.conn() as c:
        rows = c.execute(
            "SELECT s.id, s.name, COUNT(b.id) AS bill_count, "
            "COALESCE(SUM(COALESCE(b.written_total, b.computed_total, 0)), 0) AS total_spent, "
            "COALESCE(SUM(CASE WHEN b.payment_status='credit' THEN COALESCE(b.written_total, b.computed_total, 0) ELSE 0 END), 0) AS outstanding, "
            "MAX(b.bill_date) AS last_purchase "
            "FROM suppliers s LEFT JOIN bills b ON b.supplier_id = s.id "
            "AND b.status='confirmed' AND b.deleted_at IS NULL AND date(b.bill_date) >= ? AND date(b.bill_date) <= ? "
            "WHERE s.deleted_at IS NULL "
            "GROUP BY s.id ORDER BY total_spent DESC",
            (start, end),
        ).fetchall()
    return {"suppliers": [dict(r) for r in rows], "period": f"{start} to {end}"}

@router.get("/api/reports/yoy-compare")
def report_yoy_compare() -> Any:
    """Year-over-Year monthly comparison."""
    with db.conn() as c:
        current_year = datetime.now().year
        prev_year = current_year - 1
        current = c.execute(
            "SELECT strftime('%m', created_at) AS month, COALESCE(SUM(total), 0) AS revenue, COUNT(*) AS count "
            f"FROM sales WHERE strftime('%Y', created_at)=? AND {db.VALID_SALE_FILTER_NO_ALIAS} "
            "GROUP BY month ORDER BY month",
            (str(current_year),),
        ).fetchall()
        previous = c.execute(
            "SELECT strftime('%m', created_at) AS month, COALESCE(SUM(total), 0) AS revenue, COUNT(*) AS count "
            f"FROM sales WHERE strftime('%Y', created_at)=? AND {db.VALID_SALE_FILTER_NO_ALIAS} "
            "GROUP BY month ORDER BY month",
            (str(prev_year),),
        ).fetchall()
    return {"current_year": current_year, "previous_year": prev_year, "current": [dict(r) for r in current], "previous": [dict(r) for r in previous]}


# ════════════════════════════════════════════════════════════════════════════════
# v8.7 — NEW REPORTS: Profit Analysis + Sold Stock
# ════════════════════════════════════════════════════════════════════════════════

@router.get("/api/reports/profit-analysis")
def r_profit_analysis(start: str, end: str, group_by: str = "category") -> Any:
    """v8.7: Date-range profit analysis — by category (default) or by month.
    Returns per-row aggregates + totals. Excludes refunded sales."""
    return reports.profit_analysis_report(start, end, group_by)


@router.get("/api/reports/profit-analysis/export")
def r_profit_analysis_export(start: str = "", end: str = "", group_by: str = "category",
                              format: str = "csv") -> Any:
    """v8.7: Export profit analysis as CSV — formatted with title, date range,
    KPI summary block, then the detailed table + totals row.

    v8.16.6: Now supports format=pdf|excel — delegates to universal generator
    so the PDF/Excel buttons on the page produce properly branded output
    instead of always returning CSV.

    v8.16.7: `start` and `end` are now optional (default to current month range)
    — fixes 422 error when PDF button is clicked without date params.
    """
    # v8.16.7: Default to current month if start/end not provided
    if not start or not end:
        from datetime import datetime as _dt_default, timedelta as _td
        today = _dt_default.now().date()
        # Default range: last 30 days
        if not start:
            start = (today - _td(days=30)).strftime("%Y-%m-%d")
        if not end:
            end = today.strftime("%Y-%m-%d")

    # v8.16.4: For PDF/Excel, delegate to the universal generator
    if format.lower() in ("pdf", "excel"):
        data = reports.profit_analysis_report(start, end, group_by)
        from datetime import datetime as _dt3
        ts = _dt3.now().strftime("%Y%m%d_%H%M%S")
        filename = f"profit_analysis_{ts}"
        if format.lower() == "pdf":
            return _generate_pdf("profit-analysis", data, filename)
        else:
            return _generate_excel("profit-analysis", data, filename)
    import io, csv
    data = reports.profit_analysis_report(start, end, group_by)
    output = io.StringIO()
    w = csv.writer(output)

    # Title block
    w.writerow(["BillBook — Profit Analysis Report"])
    w.writerow(["Date Range", f"{start} to {end}"])
    w.writerow(["Grouped By", "Month" if group_by == "month" else "Category"])
    w.writerow([])

    # KPI summary
    t = data.get("totals", {})
    w.writerow(["SUMMARY"])
    w.writerow(["Total Revenue", t.get("revenue", 0)])
    w.writerow(["Total COGS", t.get("cogs", 0)])
    w.writerow(["Gross Profit", t.get("gross_profit", 0)])
    w.writerow(["Margin %", t.get("margin_pct", 0)])
    w.writerow(["Qty Sold", t.get("qty_sold", 0)])
    # v8.18.14: extra (non-POS) sales income — own labeled row so it's
    # differentiable from POS revenue in the export
    w.writerow(["Extra Sales Income (non-POS — cartons, raddi)", t.get("extra_sales_income", 0)])
    if group_by == "month":
        w.writerow(["Operating Expenses", t.get("operating_expenses", 0)])
        w.writerow(["Operating Profit (incl. extra sales)", t.get("operating_profit", 0)])
    w.writerow([])

    # Detailed table
    if group_by == "month":
        w.writerow(["Month", "Qty Sold", "Revenue", "COGS", "Gross Profit",
                    "Margin %", "Extra Sales (non-POS)", "Operating Expenses", "Operating Profit"])
        for m in data.get("months", []):
            w.writerow([m["month"], m["qty_sold"], m["revenue"], m["cogs"],
                        m["gross_profit"], m["margin_pct"],
                        m.get("extra_sales_income", 0),
                        m["operating_expenses"], m["operating_profit"]])
        w.writerow(["TOTAL", t.get("qty_sold", 0), t.get("revenue", 0),
                    t.get("cogs", 0), t.get("gross_profit", 0),
                    t.get("margin_pct", 0), t.get("extra_sales_income", 0),
                    t.get("operating_expenses", 0),
                    t.get("operating_profit", 0)])
    else:
        w.writerow(["Code", "Category", "Qty Sold", "Revenue", "COGS",
                    "Gross Profit", "Margin %", "Avg Selling Price", "Sale Count"])
        for cat in data.get("categories", []):
            w.writerow([cat["code"], cat["name"], cat["qty_sold"], cat["revenue"],
                        cat["cogs"], cat["gross_profit"], cat["margin_pct"],
                        cat["avg_selling_price"], cat["sale_count"]])
        w.writerow(["TOTAL", "", t.get("qty_sold", 0), t.get("revenue", 0),
                    t.get("cogs", 0), t.get("gross_profit", 0),
                    t.get("margin_pct", 0), "", ""])
    headers = {"Content-Disposition": f'attachment; filename="profit_analysis_{start}_to_{end}.csv"'}
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers=headers)


@router.get("/api/reports/sold-stock")
def r_sold_stock(start: str, end: str, group_by: str = "category") -> Any:
    """v8.7: Date-range sold stock report — by category (DEFAULT) or by item.
    Reviewer 3: 'By Category' is the default because AI-extracted item_names
    are too noisy. Excludes refunded sales."""
    return reports.sold_stock_report(start, end, group_by)


@router.get("/api/reports/sold-stock/export")
def r_sold_stock_export(start: str = "", end: str = "", group_by: str = "category",
                         format: str = "csv") -> Any:
    """v8.7: Export sold stock report as CSV.

    v8.16.4: Now supports format=pdf|excel — delegates to universal generator.

    v8.16.7: `start` and `end` are now optional (default to last 30 days)
    — fixes 422 error when PDF button is clicked without date params.
    """
    # v8.16.7: Default to last 30 days if not provided
    if not start or not end:
        from datetime import datetime as _dt_def, timedelta as _td
        today = _dt_def.now().date()
        if not start:
            start = (today - _td(days=30)).strftime("%Y-%m-%d")
        if not end:
            end = today.strftime("%Y-%m-%d")
    # v8.16.4: For PDF/Excel, delegate to the universal generator
    if format.lower() in ("pdf", "excel"):
        data = reports.sold_stock_report(start, end, group_by)
        from datetime import datetime as _dt3
        ts = _dt3.now().strftime("%Y%m%d_%H%M%S")
        filename = f"sold_stock_{ts}"
        if format.lower() == "pdf":
            return _generate_pdf("sold-stock", data, filename)
        else:
            return _generate_excel("sold-stock", data, filename)
    import io, csv
    data = reports.sold_stock_report(start, end, group_by)
    output = io.StringIO()
    w = csv.writer(output)
    if group_by == "item":
        w.writerow(["Item", "Category", "Qty Sold", "Revenue", "COGS",
                    "Gross Profit", "Margin %", "Avg Price", "Avg Cost",
                    "Sale Count", "First Sold", "Last Sold"])
        for it in data.get("items", []):
            w.writerow([it["item_name"], it["cat_name"], it["qty_sold"],
                        it["revenue"], it["cogs"], it["gross_profit"],
                        it["margin_pct"], it["avg_selling_price"],
                        it["avg_cost_price"], it["sale_count"],
                        it["first_sold"], it["last_sold"]])
        t = data.get("totals", {})
        w.writerow(["TOTAL", "", t.get("qty_sold", 0), t.get("revenue", 0),
                    t.get("cogs", 0), t.get("gross_profit", 0),
                    t.get("margin_pct", 0), "", "", "", "", ""])
    else:
        w.writerow(["Code", "Category", "Qty Sold", "Revenue", "COGS",
                    "Gross Profit", "Margin %", "Avg Price", "Sale Count",
                    "Distinct Items"])
        for cat in data.get("categories", []):
            w.writerow([cat["code"], cat["name"], cat["qty_sold"], cat["revenue"],
                        cat["cogs"], cat["gross_profit"], cat["margin_pct"],
                        cat["avg_selling_price"], cat["sale_count"],
                        cat["distinct_items"]])
        t = data.get("totals", {})
        w.writerow(["TOTAL", "", t.get("qty_sold", 0), t.get("revenue", 0),
                    t.get("cogs", 0), t.get("gross_profit", 0),
                    t.get("margin_pct", 0), "", "", t.get("distinct_categories", 0)])
    headers = {"Content-Disposition": f'attachment; filename="sold_stock_{start}_to_{end}.csv"'}
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers=headers)


# ─── v8.13.0: Category operations endpoints ────────────────────────────

@router.get("/api/reports/supplier-comparison")
def report_supplier_comparison(category_id: int = None) -> Any:
    """For each category (or one specific category), list every supplier who
    has sold you that category via confirmed bills, with their avg/last/min
    price + delta vs the shop's running avg_cost. Used to decide who to buy
    from next time.
    """
    from ..category_ops import supplier_comparison_by_category
    return {"categories": supplier_comparison_by_category(category_id)}


@router.get("/api/reports/category-cost-trends")
def report_category_cost_trends(days: int = 30, threshold_pct: float = 5.0) -> Any:
    """Per-category avg_cost trend over the last `days` days. Flags categories
    whose cost has risen > `threshold_pct`% in that window without a
    corresponding sell-price increase. Surfaces silent margin erosion.
    """
    from ..category_ops import category_cost_trend_alerts
    alerts = category_cost_trend_alerts(days=days, threshold_pct=threshold_pct)
    return {
        "days": days,
        "threshold_pct": threshold_pct,
        "alerts": alerts,
        "critical_count": sum(1 for a in alerts if a["alert_severity"] == "critical"),
        "warning_count": sum(1 for a in alerts if a["alert_severity"] == "warning"),
        "info_count": sum(1 for a in alerts if a["alert_severity"] == "info"),
    }


@router.get("/api/reports/stock-writeoffs")
def report_stock_writeoffs(month: str = "", limit: int = 200) -> Any:
    """List stock write-offs (damage / expiry / theft / sample / display)."""
    from ..category_ops import list_stock_writeoffs
    return {"writeoffs": list_stock_writeoffs(month=month, limit=limit)}


@router.get("/api/reports/stock-writeoffs/summary")
def report_stock_writeoffs_summary(month: str = "") -> Any:
    """Summary of stock write-offs for the monthly P&L 'Shrinkage' line item."""
    from ..category_ops import stock_writeoff_summary
    return stock_writeoff_summary(month=month)


# ─── v8.16.1: Universal PDF + Excel export for ALL reports ──────────────────
# Generic export endpoints that work with any report — pass the report name
# and the same query params as the original endpoint.

@router.get("/api/reports/{report_name}/export")
def universal_report_export(report_name: str, request: Request) -> Any:
    """Universal report export — supports PDF and Excel formats.
    
    Query params:
        format: "pdf" or "excel" (default: "excel")
        start: date range start
        end: date range end
        month: month filter
        ... any other params the original report endpoint accepts
    
    Usage:
        /api/reports/margins/export?format=pdf
        /api/reports/monthly/export?format=excel&start=2026-01-01&end=2026-08-31
    """
    import io, csv, json as _json
    from fastapi.responses import StreamingResponse
    from urllib.parse import urlencode
    from datetime import datetime as _dt
    
    fmt = request.query_params.get("format", "excel").lower()
    
    # Map report_name to the function that generates the data
    # v8.16.2: Expanded to cover ALL report pages in the app
    # v8.16.4: Added audit, suspicious, monthly-close, targets
    report_map = {
        "margins": ("app.profit", "get_margins"),
        "monthly": ("app.profit", "get_monthly_profit"),
        "ytd": ("app.profit", "get_ytd_profit"),
        "cash-buckets": ("app.profit_cash", "get_cash_buckets"),
        "earnings": ("app.shop", "get_actual_earnings"),
        "actual-earnings": ("app.shop", "get_actual_earnings"),
        "profit-analysis": ("app.reports", "profit_analysis_report"),
        "sold-stock": ("app.reports", "sold_stock_report"),
        "daily-stock": ("app.profit", "get_daily_stock_report"),
        "pnl": ("app.routers.reports", "pnl_report"),
        "cash-flow": ("app.routers.reports", "cash_flow_report"),
        "balance-sheet": ("app.routers.reports", "balance_sheet_report"),
        "top-items": ("app.routers.reports", "top_items_report"),
        "peak-hours": ("app.routers.reports", "peak_hours_report"),
        "ar-aging": ("app.routers.reports", "report_ar_aging"),
        "ap-aging": ("app.routers.reports", "report_ap_aging"),
        "inventory-turnover": ("app.routers.reports", "report_inventory_turnover"),
        "gmroi": ("app.routers.reports", "report_gmroi"),
        "sell-through": ("app.routers.reports", "report_sell_through"),
        "shrinkage": ("app.routers.reports", "report_shrinkage"),
        "sales-by-customer": ("app.routers.reports", "report_sales_by_customer"),
        "sales-by-employee": ("app.routers.reports", "report_sales_by_employee"),
        "atv-basket": ("app.routers.reports", "report_atv_basket"),
        "retention": ("app.routers.reports", "report_retention"),
        "supplier-performance": ("app.routers.reports", "report_supplier_performance"),
        "yoy-compare": ("app.routers.reports", "report_yoy_compare"),
        "supplier-comparison": ("app.routers.reports", "report_supplier_comparison"),
        "category-cost-trends": ("app.routers.reports", "report_category_cost_trends"),
        "stock-writeoffs": ("app.routers.reports", "report_stock_writeoffs"),
        "expenses": ("app.shop", "get_expense_summary"),
        "store-profit": ("app.profit_analytics", "get_store_profit_dashboard"),
        "overview": ("app.routers.insights", "i_dashboard"),
        "billwise": ("app.reports", "billwise_report"),
        # v8.16.4: previously missing — now wired
        "audit": ("app.auditor", "get_latest_audit_run"),
        "suspicious": ("app.shop", "list_suspicious_events"),
        "monthly-close": ("app.insights", "monthly_close"),
        "targets": ("app.pos_import", "get_target_progress"),
        # v8.18.14: Extra Sales (non-POS) report — powers the PDF/Excel/CSV
        # export buttons on the Extra Sales page
        "extra-sales": ("app.shop", "get_extra_sales_report"),
    }
    
    if report_name not in report_map:
        raise HTTPException(404, f"Unknown report: {report_name}")
    
    module_name, func_name = report_map[report_name]
    try:
        import importlib
        mod = importlib.import_module(module_name)
        func = getattr(mod, func_name)
    except (ImportError, AttributeError) as e:
        raise HTTPException(500, f"Report function not found: {e}")
    
    # Call the report function with the appropriate params
    start = request.query_params.get("start", "")
    end = request.query_params.get("end", "")
    month = request.query_params.get("month", "")
    date = request.query_params.get("date", "")
    year_param = request.query_params.get("year", "")
    group_by = request.query_params.get("group_by", "category")
    period = request.query_params.get("period", "daily")
    target_date = request.query_params.get("target_date", "")
    
    try:
        # ── v8.16.4: Special handlers for newly-wired reports ───────────
        from datetime import datetime as _dt2
        now = _dt2.now()
        if report_name == "audit":
            data = func() or {"run": None, "findings": [], "note": "No audit runs yet"}
        elif report_name == "suspicious":
            alerts_list = func(200)
            data = {"alerts": alerts_list, "count": len(alerts_list) if alerts_list else 0}
        elif report_name == "monthly-close":
            # v8.18.12 FIX: the page's month input sends month=YYYY-MM
            # (e.g. '2026-08'); int('2026-08') crashed the PDF AND Excel
            # exports with "invalid literal for int() with base 10". Accept
            # both YYYY-MM (auto-injected export buttons) and separate
            # year + month params (dashboard / command palette links).
            if month and "-" in month:
                _y, _m = month.split("-")[:2]
                y, m = int(_y), int(_m)
            else:
                y = int(year_param) if year_param else now.year
                m = int(month) if month else now.month
            if not (1 <= m <= 12 and 1900 <= y <= 2200):
                raise HTTPException(400, f"Invalid month/year: {month or f'{y}-{m}'}")
            data = func(y, m)
        elif report_name == "targets":
            td = target_date or (now.strftime("%Y-%m") if period == "monthly"
                                 else now.strftime("%Y-%m-%d"))
            data = func(period, td)
        # ── Original param-passing logic ────────────────────────────────
        # Reports that take (month) or () — no params or month only
        elif report_name in ("monthly", "ytd", "margins", "earnings", "actual-earnings", "pnl", 
                           "cash-flow", "balance-sheet", "expenses", "store-profit", "overview",
                           "extra-sales"):
            data = func(month) if month else func("")
        # Reports that take (date) — date-based
        elif report_name in ("cash-buckets", "daily-stock"):
            data = func(date) if date else func("")
        # Reports that take (start, end, group_by) — date range + grouping
        elif report_name in ("profit-analysis", "sold-stock"):
            data = func(start, end, group_by)
        # Reports that take (start, end) — date range only
        elif report_name in ("top-items", "peak-hours", "shrinkage", "sales-by-customer",
                             "sales-by-employee", "atv-basket", "supplier-performance"):
            data = func(start, end)
        # Reports that take no params
        else:
            data = func()
    except TypeError:
        # If the function doesn't accept the params we passed, try calling with no args
        try:
            data = func()
        except Exception as e:
            raise HTTPException(500, f"Report generation failed: {e}")
    except HTTPException:
        # v8.18.12: deliberate error statuses (e.g. the 400 for an invalid
        # month/year above) must pass through, not be masked as 500s.
        raise
    except Exception as e:
        raise HTTPException(500, f"Report generation failed: {e}")
    
    # Generate filename
    ts = _dt.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{report_name}_{ts}"
    
    if fmt == "pdf":
        return _generate_pdf(report_name, data, filename)
    elif fmt == "csv":
        # v8.18.12: the daily-stock page's "Export CSV" button hits this
        # route (the dedicated CSV route in profit.py is shadowed by this
        # universal route), so CSV must be supported here too.
        return _generate_csv(report_name, data, filename)
    else:
        return _generate_excel(report_name, data, filename)


def _kpi_label(key: str) -> str:
    """v8.18.14: human-friendly label for a KPI key — consults _PRETTY_COLUMNS
    (single source of truth shared with table headers) and falls back to
    title-casing. Used by _flatten_dict and _extract_kpi_groups so exported
    PDFs/Excel/CSV show 'Extra Sales Income (Non-POS)' instead of raw keys
    like 'extra_sales_income' — differentiating non-POS income from POS
    sales in every export."""
    if key in _PRETTY_COLUMNS:
        return _PRETTY_COLUMNS[key]
    return str(key).replace('_', ' ').title()


def _flatten_dict(d: dict, prefix: str = "", out: list = None, max_depth: int = 3) -> list:
    """Recursively flatten a nested dict into a list of (label, value) tuples.

    Nested keys are joined with ' — ' (e.g. "Current Margins — Actual Overall Margin").
    Lists and empty values are skipped. Stops descending at max_depth to avoid
    runaway recursion on deeply-nested dashboards.
    """
    if out is None:
        out = []
    if max_depth <= 0:
        return out
    for k, v in d.items():
        if v is None or v == "":
            continue
        label = (prefix + " — " if prefix else "") + _kpi_label(str(k))
        if isinstance(v, dict):
            _flatten_dict(v, label, out, max_depth - 1)
        elif isinstance(v, list):
            # Skip lists — they become tables elsewhere
            continue
        elif isinstance(v, bool):
            out.append((label, "Yes" if v else "No"))
        else:
            out.append((label, v))
    return out


def _extract_kpi_groups(data: dict) -> list:
    """Extract KPIs from a report dict, grouping nested-dict sections.

    Returns a list of (section_name, [(label, value), ...]) tuples.
    Section name is "" for top-level scalars; otherwise the parent key name.
    Skips metadata keys (note, report_name, etc.) and empty values.
    """
    META_KEYS = {"note", "report_name", "period_start", "period_end",
                 "currency", "as_of", "date", "month", "today",
                 "opening_date", "warning", "missing_categories_warning",
                 "missing_categories", "threshold_pct", "days", "source"}

    groups: list = []
    flat_kpis: list = []

    for key, val in data.items():
        if key in META_KEYS:
            continue
        if val is None or val == "":
            continue
        if isinstance(val, bool):
            flat_kpis.append((_kpi_label(str(key)),
                              "Yes" if val else "No"))
        elif isinstance(val, (int, float, str)):
            flat_kpis.append((_kpi_label(str(key)), val))
        elif isinstance(val, dict):
            # Recursively flatten this section
            section_name = _kpi_label(str(key))
            section_kpis = _flatten_dict(val, "", [], max_depth=3)
            if section_kpis:
                groups.append((section_name, section_kpis))
        elif isinstance(val, list):
            # Skip lists — they become tables
            continue

    # Put flat KPIs first as the "Summary" section
    if flat_kpis:
        groups.insert(0, ("Summary", flat_kpis))
    return groups


def _find_all_tables(data: dict) -> list:
    """Find ALL lists-of-dicts in the report data.

    Returns a list of (table_title, rows) tuples. Used for rendering
    multiple data tables in a single report (e.g. store-profit has
    per_category, buckets, etc.).
    """
    tables: list = []
    if not isinstance(data, dict):
        return tables
    for key, val in data.items():
        if not isinstance(val, list):
            continue
        if len(val) == 0:
            continue
        # v8.16.12: Apply column renaming so PDF/Excel headers match UI
        first = val[0]
        if isinstance(first, dict):
            title = str(key).replace('_', ' ').title()
            # Apply pretty column names
            renamed_rows = [_rename_columns(row) for row in val]
            tables.append((title, renamed_rows))
        elif isinstance(first, (int, float, str)):
            # Simple list — render as single-column
            title = str(key).replace('_', ' ').title()
            rows = [{"value": v} for v in val]
            tables.append((title, rows))
    return tables


# v8.16.12: Maps internal field names to human-friendly display names that match the UI.
# Both PDF and Excel generators use this so column headers are consistent across all 3.
_PRETTY_COLUMNS = {
    # Common fields
    "category_id": "Category ID",
    "code": "Code",
    "name": "Category",
    "sell_price": "Sell Price",
    "qty_sold": "Qty Sold",
    "revenue": "Revenue",
    "cogs": "COGS",
    "gross_profit": "Gross Profit",
    "sale_count": "Sales",
    # Profit Analysis fields
    "margin_pct": "Hist. Margin %",
    "current_avg_cost": "Curr. Avg Cost",
    "current_margin_pct": "Curr. Margin %",
    "avg_historical_cost": "Avg Hist. Cost",
    "cost_change": "Cost Δ",
    "avg_selling_price": "Avg Selling Price",
    "profit_per_unit": "Profit/Unit",
    "current_profit_per_unit": "Curr. Profit/Unit",
    "markup_pct": "Markup %",
    "current_markup_pct": "Curr. Markup %",
    "margin_per_unit": "Margin/Unit",
    # v8.18.14: Extra (non-POS) Sales fields — every appearance of these
    # keys in KPI sections or table columns gets an explicit "(Non-POS)"
    # tag so non-POS income is always differentiable from POS sales.
    "extra_sales_income": "Extra Sales Income (Non-POS)",
    "extra_sales_total": "Extra Sales Total (Non-POS)",
    "extra_sales_count": "Extra Sales Entries",
    "extra_sales_cash": "Extra Sales — Cash (Non-POS)",
    "extra_sales_other": "Extra Sales — Bank/Card (Non-POS)",
    "ytd_extra_sales_income": "YTD Extra Sales Income (Non-POS)",
    "other_income": "Other Income (Extra Sales, Non-POS)",
}


def _rename_columns(row: dict) -> dict:
    """Rename a row's keys from internal names to display names.

    Used by both Excel and PDF generators so column headers match what the UI shows.
    Unknown keys are kept as-is (title-cased).
    """
    if not isinstance(row, dict):
        return row
    out = {}
    for k, v in row.items():
        if k in _PRETTY_COLUMNS:
            out[_PRETTY_COLUMNS[k]] = v
        else:
            # Title-case the key as a fallback
            out[str(k).replace('_', ' ').title()] = v
    return out


def _generate_excel(report_name: str, data: dict, filename: str) -> Any:
    """Generate professional Excel (.xlsx) from report data.

    v8.16.5: Now properly handles nested dashboard data — recursively flattens
    nested dicts into KPI cards (grouped by section), and renders ALL lists
    of dicts as separate labeled tables (not just the first one).
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    # ── Brand palette (matches the app's cream/copper theme) ──────────────────
    C_BRAND   = "1F3A5F"   # deep navy
    C_ACCENT  = "CC785C"   # copper
    C_CREAM   = "FAF9F5"
    C_STRIPE  = "F5F0E8"
    C_BORDER  = "E6DFD8"
    C_KPI_BG  = "EDF1F6"
    C_SECTION = "E8E2D5"   # section header background

    thin = Side(border_style="thin", color=C_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = report_name[:31]  # Excel sheet name max 31 chars
    ws.sheet_view.showGridLines = False

    pretty_name = report_name.replace('-', ' ').title()
    now_str = _dt_now()

    # ── Title band (rows 1-2) ────────────────────────────────────────────────
    n_cols = 6
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    t = ws.cell(row=1, column=1, value=f"BillBook — {pretty_name}")
    t.font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    t.fill = PatternFill("solid", fgColor=C_BRAND)
    ws.row_dimensions[1].height = 30

    s = ws.cell(row=2, column=1, value=f"Generated: {now_str}    |    BillBook POS + Billing System")
    s.font = Font(name="Calibri", size=10, italic=True, color="FFFFFF")
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    s.fill = PatternFill("solid", fgColor=C_ACCENT)
    ws.row_dimensions[2].height = 18

    row_num = 4

    # ── KPI summary band — grouped by section ─────────────────────────────────
    if isinstance(data, dict):
        kpi_groups = _extract_kpi_groups(data)
        for section_name, kpi_pairs in kpi_groups:
            if not kpi_pairs:
                continue
            # Section heading
            sh = ws.cell(row=row_num, column=1, value=section_name.upper())
            sh.font = Font(name="Calibri", size=10, bold=True, color=C_BRAND)
            sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            sh.fill = PatternFill("solid", fgColor=C_SECTION)
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=n_cols)
            ws.row_dimensions[row_num].height = 18
            row_num += 1

            # Render KPI cards: 2 per row, label+value pairs
            for i in range(0, len(kpi_pairs), 2):
                chunk = kpi_pairs[i:i+2]
                for j, (k, v) in enumerate(chunk):
                    col_label = 1 + j*3
                    # Label
                    lc = ws.cell(row=row_num, column=col_label, value=k)
                    lc.font = Font(name="Calibri", size=9, bold=True, color="6B7280")
                    lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                    lc.fill = PatternFill("solid", fgColor=C_KPI_BG)
                    ws.merge_cells(start_row=row_num, start_column=col_label,
                                   end_row=row_num, end_column=col_label+1)
                    # Value
                    vc = ws.cell(row=row_num+1, column=col_label, value=v)
                    vc.font = Font(name="Calibri", size=12, bold=True, color=C_BRAND)
                    vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                    vc.fill = PatternFill("solid", fgColor=C_KPI_BG)
                    ws.merge_cells(start_row=row_num+1, start_column=col_label,
                                   end_row=row_num+1, end_column=col_label+1)
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        vc.number_format = '#,##0.00' if isinstance(v, float) else '#,##0'
                ws.row_dimensions[row_num].height = 16
                ws.row_dimensions[row_num+1].height = 22
                row_num += 3
            row_num += 1  # gap between sections

    # ── All data tables — render every list-of-dicts ──────────────────────────
    if isinstance(data, dict):
        tables = _find_all_tables(data)
        for table_title, rows in tables:
            if not rows:
                continue
            cols = list(rows[0].keys())
            pretty_cols = [str(c).replace('_', ' ').title() for c in cols]

            # Table heading
            th = ws.cell(row=row_num, column=1, value=f"📊 {table_title}")
            th.font = Font(name="Calibri", size=11, bold=True, color=C_BRAND)
            th.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=max(len(cols), 4))
            ws.row_dimensions[row_num].height = 20
            row_num += 1

            # Header row
            for ci, (col, pcol) in enumerate(zip(cols, pretty_cols), 1):
                c = ws.cell(row=row_num, column=ci, value=pcol)
                c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=C_BRAND)
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                c.border = border
            ws.row_dimensions[row_num].height = 22
            row_num += 1

            # Data rows
            for ri, row in enumerate(rows):
                stripe = C_STRIPE if ri % 2 else C_CREAM
                for ci, col in enumerate(cols, 1):
                    val = row.get(col, "")
                    if isinstance(val, bool):
                        val = "Yes" if val else "No"
                    elif isinstance(val, dict):
                        val = ", ".join(f"{k}={v}" for k, v in val.items() if v is not None)
                    elif isinstance(val, list):
                        val = ", ".join(str(x) for x in val)
                    c = ws.cell(row=row_num, column=ci, value=val if val is not None else "")
                    c.font = Font(name="Calibri", size=10, color="1F2937")
                    c.fill = PatternFill("solid", fgColor=stripe)
                    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                    c.border = border
                    if isinstance(val, (int, float)) and not isinstance(val, bool):
                        c.number_format = '#,##0.00' if isinstance(val, float) else '#,##0'
                ws.row_dimensions[row_num].height = 18
                row_num += 1

            # Auto-width columns based on this table
            for ci, col in enumerate(cols, 1):
                max_len = len(pretty_cols[ci-1])
                for row in rows[:200]:
                    v = row.get(col, "")
                    if v is None:
                        continue
                    s = str(v)
                    if len(s) > max_len:
                        max_len = len(s)
                current = ws.column_dimensions[get_column_letter(ci)].width or 10
                ws.column_dimensions[get_column_letter(ci)].width = max(current, min(max_len + 4, 40))

            # Footer caption
            cap = ws.cell(row=row_num, column=1,
                          value=f"   {len(rows)} records")
            cap.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=max(len(cols), 4))
            row_num += 2  # gap between tables

    # ── Footer band ───────────────────────────────────────────────────────────
    row_num += 1
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=n_cols)
    f = ws.cell(row=row_num, column=1,
                value="Generated by BillBook  •  Self-hosted POS + Billing + Business Management")
    f.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    f.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    f.fill = PatternFill("solid", fgColor=C_CREAM)
    ws.row_dimensions[row_num].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
    )


def _generate_pdf(report_name: str, data: dict, filename: str) -> Any:
    """Generate professional PDF from report data.

    v8.16.5: Now properly handles nested dashboard data — recursively flattens
    nested dicts into KPI cards (grouped by section), and renders ALL lists
    of dicts as separate labeled tables (not just the first one).
    """
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    from reportlab.platypus import (Table, TableStyle, Paragraph, Spacer,
                                     Frame, PageTemplate, BaseDocTemplate,
                                     KeepTogether)
    from reportlab.platypus.flowables import HRFlowable
    from fastapi.responses import StreamingResponse

    # Brand palette
    C_BRAND   = colors.HexColor('#1F3A5F')
    C_ACCENT  = colors.HexColor('#CC785C')
    C_CREAM   = colors.HexColor('#FAF9F5')
    C_STRIPE  = colors.HexColor('#F5F0E8')
    C_BORDER  = colors.HexColor('#E6DFD8')
    C_DIM     = colors.HexColor('#6B7280')
    C_TEXT    = colors.HexColor('#1F2937')
    C_SECTION = colors.HexColor('#E8E2D5')

    pretty_name = report_name.replace('-', ' ').title()
    now_str = _dt_now()

    buf = io.BytesIO()
    page_size = landscape(A4)
    page_w, page_h = page_size

    # ── Page template with header + footer ────────────────────────────────────
    def _draw_page(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(C_ACCENT)
        canvas.rect(0, page_h - 4*mm, page_w, 4*mm, stroke=0, fill=1)
        canvas.setFillColor(C_BRAND)
        canvas.setFont('Helvetica-Bold', 9)
        canvas.drawString(15*mm, page_h - 12*mm, "BillBook")
        canvas.setFillColor(C_DIM)
        canvas.setFont('Helvetica-Oblique', 8)
        canvas.drawRightString(page_w - 15*mm, page_h - 12*mm,
                               f"{pretty_name}  •  {now_str}")
        canvas.setFillColor(C_DIM)
        canvas.setFont('Helvetica', 8)
        canvas.drawString(15*mm, 8*mm,
                          "Self-hosted POS + Billing + Business Management")
        canvas.drawRightString(page_w - 15*mm, 8*mm,
                              f"Page {canvas.getPageNumber()}")
        canvas.restoreState()

    doc = BaseDocTemplate(
        buf, pagesize=page_size,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=18*mm, bottomMargin=14*mm,
        title=f"BillBook — {pretty_name}",
        author="BillBook",
    )
    frame = Frame(15*mm, 14*mm, page_w - 30*mm, page_h - 32*mm,
                  leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0)
    doc.addPageTemplates([PageTemplate(id='all', frames=[frame], onPage=_draw_page)])

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('BrandTitle', parent=styles['Title'],
                                  fontName='Helvetica-Bold', fontSize=20,
                                  textColor=C_BRAND, alignment=TA_LEFT,
                                  spaceAfter=2, leading=24)
    sub_style = ParagraphStyle('BrandSub', parent=styles['Normal'],
                               fontName='Helvetica', fontSize=9,
                               textColor=C_DIM, alignment=TA_LEFT,
                               spaceAfter=8, leading=12)
    section_style = ParagraphStyle('Section', parent=styles['Normal'],
                                   fontName='Helvetica-Bold', fontSize=10,
                                   textColor=C_BRAND, alignment=TA_LEFT,
                                   spaceBefore=8, spaceAfter=4, leading=12)
    kpi_label_style = ParagraphStyle('KPILabel', parent=styles['Normal'],
                                     fontName='Helvetica-Bold', fontSize=7,
                                     textColor=C_DIM, alignment=TA_LEFT, leading=9)
    kpi_value_style = ParagraphStyle('KPIValue', parent=styles['Normal'],
                                     fontName='Helvetica-Bold', fontSize=11,
                                     textColor=C_BRAND, alignment=TA_LEFT, leading=13)
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'],
                                fontName='Helvetica', fontSize=7.5,
                                textColor=C_TEXT, alignment=TA_LEFT, leading=9)
    table_title_style = ParagraphStyle('TableTitle', parent=styles['Normal'],
                                        fontName='Helvetica-Bold', fontSize=11,
                                        textColor=C_BRAND, alignment=TA_LEFT,
                                        spaceBefore=10, spaceAfter=4, leading=14)
    note_style = ParagraphStyle('Note', fontName='Helvetica-Oblique', fontSize=8,
                                  textColor=C_DIM)
    elements = []

    # ── Title block ────────────────────────────────────────────────────────────
    elements.append(Paragraph(pretty_name, title_style))
    elements.append(Paragraph(
        f"BillBook Report  •  Generated {now_str}", sub_style))
    elements.append(HRFlowable(width="100%", thickness=0.6,
                               color=C_ACCENT, spaceBefore=2, spaceAfter=8))

    # ── KPI summary band — grouped by section ─────────────────────────────────
    if isinstance(data, dict):
        kpi_groups = _extract_kpi_groups(data)
        # Section layout: 4 KPI cards per row
        available_width = page_w - 30*mm
        card_width = (available_width - 12*mm) / 4  # 4 cards with 4mm gaps

        for section_name, kpi_pairs in kpi_groups:
            if not kpi_pairs:
                continue
            section_flowables = [Paragraph(section_name.upper(), section_style)]
            # Build KPI card rows (4 cards per row)
            for i in range(0, len(kpi_pairs), 4):
                chunk = kpi_pairs[i:i+4]
                cells = []
                for k, v in chunk:
                    val_str = (f"{v:,.2f}" if isinstance(v, float)
                               else f"{v:,}" if isinstance(v, int) and not isinstance(v, bool)
                               else str(v))
                    if len(val_str) > 22:
                        val_str = val_str[:19] + "..."
                    inner = Table(
                        [[Paragraph(k.upper(), kpi_label_style)],
                         [Paragraph(val_str, kpi_value_style)]],
                        colWidths=[card_width], rowHeights=[5*mm, 7*mm]
                    )
                    inner.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), C_CREAM),
                        ('LEFTPADDING', (0, 0), (-1, -1), 5),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                        ('TOPPADDING', (0, 0), (-1, -1), 2),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                        ('LINEBEFORE', (0, 0), (0, -1), 2, C_ACCENT),
                        ('BOX', (0, 0), (-1, -1), 0.3, C_BORDER),
                    ]))
                    cells.append(inner)
                while len(cells) < 4:
                    cells.append("")
                row = Table([cells], colWidths=[card_width + 3*mm]*4,
                            rowHeights=[13*mm])
                row.setStyle(TableStyle([
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                section_flowables.append(row)
            # Try to keep each section together
            elements.append(KeepTogether(section_flowables))
            elements.append(Spacer(1, 4))

    # ── All data tables — render every list-of-dicts ──────────────────────────
    if isinstance(data, dict):
        tables = _find_all_tables(data)
        for table_title, rows in tables:
            if not rows:
                continue
            # Truncate rows for PDF (max 50 per table to fit on a page)
            display_rows = rows[:50]
            cols = list(display_rows[0].keys())
            pretty_cols = [str(c).replace('_', ' ').title() for c in cols]

            # Header row
            table_data = [[Paragraph(c, ParagraphStyle('HDR',
                                                      fontName='Helvetica-Bold',
                                                      fontSize=8, textColor=colors.white,
                                                      leading=10))
                            for c in pretty_cols]]
            # Data rows
            for row in display_rows:
                row_cells = []
                for c in cols:
                    v = row.get(c, "")
                    if v is None:
                        s = ""
                    elif isinstance(v, bool):
                        s = "Yes" if v else "No"
                    elif isinstance(v, dict):
                        s = ", ".join(f"{k}={val}" for k, val in v.items() if val is not None)
                        if len(s) > 50:
                            s = s[:47] + "..."
                    elif isinstance(v, list):
                        s = ", ".join(str(x) for x in v)
                        if len(s) > 50:
                            s = s[:47] + "..."
                    elif isinstance(v, float):
                        s = f"{v:,.2f}"
                    elif isinstance(v, int):
                        s = f"{v:,}"
                    else:
                        s = str(v)
                    if len(s) > 60:
                        s = s[:57] + "..."
                    row_cells.append(Paragraph(s, cell_style))
                table_data.append(row_cells)

            # Column widths
            n_cols = len(cols)
            base_width = available_width / n_cols
            col_widths = [min(base_width, 60*mm) for _ in cols]
            scale = available_width / sum(col_widths)
            col_widths = [w * scale for w in col_widths]

            t = Table(table_data, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), C_BRAND),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 1), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [C_CREAM, C_STRIPE]),
                ('GRID', (0, 0), (-1, -1), 0.3, C_BORDER),
                ('LINEBELOW', (0, 0), (-1, 0), 1, C_ACCENT),
            ]))

            # Wrap title + table together so they don't get split across pages
            table_block = [
                Paragraph(f"📊 {table_title}", table_title_style),
                t,
            ]
            if len(rows) > 50:
                table_block.append(Paragraph(
                    f"<i>Showing 50 of {len(rows)} rows — export as Excel for full data.</i>",
                    note_style))
            elements.append(KeepTogether(table_block))
            elements.append(Spacer(1, 4))

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}.pdf"}
    )


def _dt_now():
    from datetime import datetime
    return datetime.now().strftime("%Y-%m-%d %H:%M")


def _csv_cell(v):
    """Flatten one cell value for CSV output (dicts/lists -> readable text)."""
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, dict):
        return ", ".join(f"{k}={x}" for k, x in v.items() if x is not None)
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    return "" if v is None else v


def _generate_csv(report_name: str, data: dict, filename: str) -> Any:
    """v8.18.12: CSV output for the universal export route.

    Mirrors the KPI-sections + tables structure of the PDF/Excel generators
    so all three formats of a report carry the same content. Needed because
    the daily-stock page's \"Export CSV\" button targets this route, and the
    dedicated CSV route in profit.py is shadowed by this universal route.
    """
    import io
    import csv as _csv
    from fastapi.responses import StreamingResponse

    out = io.StringIO()
    out.write("\ufeff")  # BOM so Excel auto-detects UTF-8
    w = _csv.writer(out)
    pretty_name = report_name.replace('-', ' ').title()
    w.writerow([f"BillBook — {pretty_name}"])
    w.writerow([f"Generated: {_dt_now()}", "BillBook POS + Billing System"])

    if isinstance(data, dict):
        for section_name, kpi_pairs in _extract_kpi_groups(data):
            if not kpi_pairs:
                continue
            w.writerow([])
            w.writerow([section_name.upper()])
            for k, v in kpi_pairs:
                w.writerow([k, _csv_cell(v)])
        for table_title, rows in _find_all_tables(data):
            if not rows:
                continue
            w.writerow([])
            w.writerow([f"TABLE: {table_title}"])
            cols = list(rows[0].keys())
            w.writerow(cols)
            for row in rows:
                w.writerow([_csv_cell(row.get(c, "")) for c in cols])
            w.writerow([f"{len(rows)} records"])
    else:
        w.writerow([])
        w.writerow([_csv_cell(data)])

    headers = {"Content-Disposition": f'attachment; filename="{filename}.csv"'}
    return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
                             headers=headers)
