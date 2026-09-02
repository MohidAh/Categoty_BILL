"""v8.18.14 — Extra Sales (non-POS) visibility + differentiability in reports.

Audit result: v8.18.13 added extra sales to P&L, Actual Earnings, Cash Flow
and the Daily Summary — but NOT to five other reports. This suite locks in
the fixes:

  1. Monthly Profit      — extra_sales_income line, operating profit includes it
  2. YTD Profit          — ytd_extra_sales_income, per-month column
  3. Store Profit        — monthly + ytd sections expose the extra income
  4. Monthly Close       — extra sales keys + labeled details rows + net profit math
                           + dedicated monthly-close.pdf renders the sell side
  5. Profit Analysis     — month mode: own column + op-profit math;
                           category mode: top-level total, never in category rows
  6. Export labels       — universal PDF/Excel/CSV KPI labels say
                           "Extra Sales Income (Non-POS)" (differentiability)
  7. New report          — /api/reports/extra-sales/export (PDF + Excel)
                           wired to the Extra Sales page buttons
  8. CSV export          — profit-analysis CSV has the Extra Sales column

Run: .venv/bin/python -m pytest tests/test_v8_18_14_extra_sales_reports.py -v
"""
import io
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "tests"))

from test_helpers import setup_test_db_with_password, cleanup, login_client

from datetime import datetime as _dt

NOW = _dt.now()
TODAY = NOW.strftime("%Y-%m-%d")
THIS_MONTH = NOW.strftime("%Y-%m")
THIS_YEAR, THIS_MONTH_NUM = NOW.year, NOW.month
M_START = f"{THIS_MONTH}-01"
M_END = f"{THIS_MONTH}-28"

JS = PROJECT_ROOT / "app" / "static" / "js"


def setup_test_db():
    return setup_test_db_with_password(prefix="billbook_v81814_")


def _seed_extra_sales(tc, month=None):
    """Two extra sales: 2000 cash + 1000 bank = 3000 total for the month.

    Both are dated TODAY: the YTD window is (opening_date .. today) and a
    fresh test DB with no POS sales gets opening_date = today — so anything
    dated even one day earlier would fall OUTSIDE the YTD range.
    """
    r1 = tc.post("/api/extra-sales", json={
        "item_name": "Cardboard cartons", "quantity": 40, "unit_price": 50,
        "payment_method": "cash", "date": TODAY})
    r2 = tc.post("/api/extra-sales", json={
        "item_name": "Raddi (scrap)", "quantity": 10, "unit_price": 100,
        "payment_method": "bank", "date": TODAY})
    assert r1.status_code == 200 and r2.status_code == 200, (r1.text, r2.text)


def _pdf_text(body: bytes) -> str:
    """Extract text from a ReportLab PDF.

    ReportLab chains /ASCII85Decode + /FlateDecode on page content streams,
    so: strip the trailing '~>', base85-decode, then zlib-inflate.
    """
    import base64
    import re
    import zlib
    parts = []
    for m in re.finditer(rb"stream\r?\n", body):
        start = m.end()
        end = body.find(b"endstream", start)
        if end == -1:
            continue
        chunk = body[start:end].strip()
        try:
            a = chunk
            if a.endswith(b"~>"):
                a = a[:-2]
            if a.startswith(b"<~"):
                a = a[2:]
            parts.append(zlib.decompress(base64.a85decode(a, ignorechars=b" \t\n\r\v")))
        except Exception:
            try:
                parts.append(zlib.decompress(chunk))
            except Exception:
                parts.append(chunk)  # uncompressed stream
    return b"\n".join(parts).decode("latin-1", errors="ignore")


# ------------------------------------------------------------------
# 1. Monthly Profit
# ------------------------------------------------------------------

def test_monthly_profit_includes_extra_sales():
    d = setup_test_db()
    try:
        from fastapi.testclient import TestClient
        from app.main import app
        from app import profit_analytics
        with TestClient(app) as tc:
            login_client(tc)
            sales_before = profit_analytics.get_monthly_profit(THIS_MONTH)["sales"]
            _seed_extra_sales(tc)
            m = profit_analytics.get_monthly_profit(THIS_MONTH)
            # own line, never merged into POS sales
            assert m["extra_sales_income"] == 3000.0
            assert m["sales"] == sales_before
            # operating profit = gross profit + extra - op exp
            assert m["operating_profit"] == round(
                m["gross_profit"] + 3000.0 - m["operating_expenses"], 2)
            assert "Extra Sales" in m["note"]
    finally:
        cleanup(d)


# ------------------------------------------------------------------
# 2. YTD Profit
# ------------------------------------------------------------------

def test_ytd_profit_includes_extra_sales():
    d = setup_test_db()
    try:
        from fastapi.testclient import TestClient
        from app.main import app
        from app import profit_analytics
        with TestClient(app) as tc:
            login_client(tc)
            ytd_sales_before = profit_analytics.get_ytd_profit()["ytd_sales"]
            _seed_extra_sales(tc)
            y = profit_analytics.get_ytd_profit()
            assert y["ytd_extra_sales_income"] == 3000.0
            assert y["ytd_sales"] == ytd_sales_before  # never merged into POS sales
            assert y["ytd_operating_profit"] == round(
                y["ytd_gross_profit"] + 3000.0 - y["ytd_operating_expenses"], 2)
            # monthly rows carry the extra income column
            m09 = [m for m in y["monthly"] if m["month"] == THIS_MONTH]
            assert m09 and m09[0]["extra_sales_income"] == 3000.0
    finally:
        cleanup(d)


# ------------------------------------------------------------------
# 3. Store Profit Dashboard
# ------------------------------------------------------------------

def test_store_profit_dashboard_sections_expose_extra():
    d = setup_test_db()
    try:
        from fastapi.testclient import TestClient
        from app.main import app
        from app import profit_analytics
        with TestClient(app) as tc:
            login_client(tc)
            _seed_extra_sales(tc)
            dash = profit_analytics.get_store_profit_dashboard()
            assert dash["monthly"]["extra_sales_income"] == 3000.0
            assert dash["ytd"]["ytd_extra_sales_income"] == 3000.0
            y = profit_analytics.get_ytd_profit()
            assert dash["ytd"]["ytd_operating_profit"] == y["ytd_operating_profit"]
            assert y["ytd_operating_profit"] == round(
                y["ytd_gross_profit"] + 3000.0 - y["ytd_operating_expenses"], 2)
    finally:
        cleanup(d)


# ------------------------------------------------------------------
# 4. Monthly Close (data + dedicated PDF)
# ------------------------------------------------------------------

def test_monthly_close_includes_extra_sales():
    d = setup_test_db()
    try:
        from fastapi.testclient import TestClient
        from app.main import app
        from app import insights
        with TestClient(app) as tc:
            login_client(tc)
            _seed_extra_sales(tc)
            r = insights.monthly_close(THIS_YEAR, THIS_MONTH_NUM)
            assert r["extra_sales_count"] == 2
            assert r["extra_sales_income"] == 3000.0
            assert r["net_profit"] == round(
                r["gross_profit"] + 3000.0 - r["operating_expenses"], 2)
            # line items list is exposed for exports
            assert len(r["extra_sales"]) == 2
            assert r["extra_sales"][0]["item_name"] in ("Cardboard cartons", "Raddi (scrap)")
            # details rows carry differentiating labels
            assert "Extra Sales (non-POS" in " ".join(r["details"].keys())
            assert "Net Profit (gross + extra sales" in " ".join(r["details"].keys())
    finally:
        cleanup(d)


def test_monthly_close_pdf_renders_extra_sales():
    d = setup_test_db()
    try:
        from fastapi.testclient import TestClient
        from app.main import app
        with TestClient(app) as tc:
            login_client(tc)
            _seed_extra_sales(tc)
            r = tc.get(f"/api/reports/monthly-close.pdf?year={THIS_YEAR}&month={THIS_MONTH_NUM}")
            assert r.status_code == 200, r.text
            assert r.headers["content-type"].startswith("application/pdf")
            body = r.content
            assert body[:4] == b"%PDF"
            # decompress all content streams and look for the labels
            text = _pdf_text(body)
            assert "Extra Sales" in text, "PDF does not mention Extra Sales"
            assert "Sales & Income Summary" in text.replace("&amp;", "&"), \
                "PDF is missing the sell-side summary section"
    finally:
        cleanup(d)


# ------------------------------------------------------------------
# 5. Profit Analysis (month + category modes)
# ------------------------------------------------------------------

def test_profit_analysis_month_mode_extra_sales_column():
    d = setup_test_db()
    try:
        from fastapi.testclient import TestClient
        from app.main import app
        from app import reports
        with TestClient(app) as tc:
            login_client(tc)
            _seed_extra_sales(tc)
            r = reports.profit_analysis_report(M_START, M_END, "month")
            # months with ONLY extra sales still appear (v8.18.14) — the
            # income never silently vanishes from the table
            m09 = [m for m in r["months"] if m["month"] == THIS_MONTH]
            assert m09, "extra-sales-only month missing from month rows"
            assert m09[0]["extra_sales_income"] == 3000.0
            assert m09[0]["revenue"] == 0.0  # POS revenue untouched
            assert m09[0]["operating_profit"] == 3000.0
            assert r["totals"]["extra_sales_income"] == 3000.0
            # add a POS sale so the month row has both income streams
            r_sale = tc.post("/api/sales", json={
                "customer_name": "Cash Walk-in",
                "items": [{"category_id": 1, "category_code": "A", "sell_price": 500, "qty": 1}],
                "payment_method": "cash"})
            assert r_sale.status_code == 200, r_sale.text
            r2 = reports.profit_analysis_report(M_START, M_END, "month")
            m09 = [m for m in r2["months"] if m["month"] == THIS_MONTH][0]
            assert m09["revenue"] == 500.0
            assert m09["extra_sales_income"] == 3000.0
            assert m09["operating_profit"] == round(
                m09["gross_profit"] + 3000.0 - m09["operating_expenses"], 2)
    finally:
        cleanup(d)


def test_profit_analysis_category_mode_top_level_extra():
    d = setup_test_db()
    try:
        from fastapi.testclient import TestClient
        from app.main import app
        from app import reports
        with TestClient(app) as tc:
            login_client(tc)
            _seed_extra_sales(tc)
            r = reports.profit_analysis_report(M_START, M_END, "category")
            # top-level total exists and is NOT baked into any category row
            assert r["extra_sales_income"] == 3000.0
            assert r["totals"]["extra_sales_income"] == 3000.0
            for cat in r.get("categories", []):
                assert cat.get("revenue", 0) == 0  # extra sales never inflate revenue
    finally:
        cleanup(d)


# ------------------------------------------------------------------
# 6. Universal export labels (differentiability)
# ------------------------------------------------------------------

def test_kpi_labels_pretty_for_extra_sales_fields():
    from app.routers import reports as rr
    # _flatten_dict consults _PRETTY_COLUMNS via _kpi_label
    flat = rr._flatten_dict({"extra_sales_income": 3000.0, "other_income": 1.0})
    labels = [l for l, _ in flat]
    assert "Extra Sales Income (Non-POS)" in labels
    assert "Other Income (Extra Sales, Non-POS)" in labels
    # _extract_kpi_groups labels top-level + nested sections too
    groups = dict(rr._extract_kpi_groups({
        "extra_sales_income": 3000.0,
        "inflows": {"extra_sales_cash": 2000.0, "extra_sales_other": 1000.0},
        "ytd": {"ytd_extra_sales_income": 3000.0},
    }))
    flat_kpis = [l for l, _ in groups.get("Summary", [])]
    assert "Extra Sales Income (Non-POS)" in flat_kpis
    inflow_labels = [l for l, _ in groups.get("Inflows", [])]
    assert "Extra Sales — Cash (Non-POS)" in inflow_labels
    assert "Extra Sales — Bank/Card (Non-POS)" in inflow_labels
    ytd_labels = [l for l, _ in groups.get("Ytd", [])]
    assert "YTD Extra Sales Income (Non-POS)" in ytd_labels


def test_excel_export_carries_extra_sales_labels():
    d = setup_test_db()
    try:
        from fastapi.testclient import TestClient
        from app.main import app
        with TestClient(app) as tc:
            login_client(tc)
            _seed_extra_sales(tc)
            r = tc.get(f"/api/reports/monthly/export?format=excel&month={THIS_MONTH}")
            assert r.status_code == 200, r.text
            assert r.headers["content-type"].startswith(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            assert r.content[:2] == b"PK"
            # open the workbook and look for the differentiating label
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(r.content))
            texts = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    texts.extend(str(v) for v in row if v is not None)
            joined = " | ".join(texts)
            assert "Extra Sales Income (Non-POS)" in joined
    finally:
        cleanup(d)


# ------------------------------------------------------------------
# 7. New extra-sales report export (page buttons)
# ------------------------------------------------------------------

def test_extra_sales_report_function_and_exports():
    d = setup_test_db()
    try:
        from fastapi.testclient import TestClient
        from app.main import app
        from app import shop
        with TestClient(app) as tc:
            login_client(tc)
            _seed_extra_sales(tc)
            rep = shop.get_extra_sales_report(THIS_MONTH)
            assert rep["total_income"] == 3000.0
            assert rep["entries"] == 2
            assert len(rep["sales_list"]) == 2
            assert rep["report_title"].startswith("Extra Sales")
            # PDF export
            r = tc.get(f"/api/reports/extra-sales/export?format=pdf&month={THIS_MONTH}")
            assert r.status_code == 200, r.text
            assert r.content[:4] == b"%PDF"
            # Excel export — check it opens and carries the entries table
            r2 = tc.get(f"/api/reports/extra-sales/export?format=excel&month={THIS_MONTH}")
            assert r2.status_code == 200, r2.text
            assert r2.content[:2] == b"PK"
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(r2.content))
            texts = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    texts.extend(str(v) for v in row if v is not None)
            joined = " | ".join(texts)
            assert "Cardboard cartons" in joined
            # sales_list table columns (qty 40 x rate 50 = total 2000)
            assert "2000" in joined
            assert "Raddi (scrap)" in joined
            # CSV export (same universal route)
            r3 = tc.get(f"/api/reports/extra-sales/export?format=csv&month={THIS_MONTH}")
            assert r3.status_code == 200
            assert b"Cardboard cartons" in r3.content
    finally:
        cleanup(d)


# ------------------------------------------------------------------
# 8. Profit Analysis CSV has the Extra Sales column
# ------------------------------------------------------------------

def test_profit_analysis_csv_extra_sales_column():
    d = setup_test_db()
    try:
        from fastapi.testclient import TestClient
        from app.main import app
        with TestClient(app) as tc:
            login_client(tc)
            _seed_extra_sales(tc)
            r = tc.get(f"/api/reports/profit-analysis/export?start={M_START}&end={M_END}&group_by=month")
            assert r.status_code == 200
            body = r.content.decode("utf-8")
            assert "Extra Sales" in body
            assert "Extra Sales Income (non-POS" in body
    finally:
        cleanup(d)


# ------------------------------------------------------------------
# 9. Static JS contract — pages actually render the labels
# ------------------------------------------------------------------

def test_js_pages_render_extra_sales_labels():
    checks = [
        (JS / "pages" / "monthly-profit-page.js",
         ["+ Extra Sales (non-POS", "Extra Sales:"]),
        (JS / "pages" / "ytd-profit-page.js",
         ["YTD Extra Sales (non-POS)", "ytd_extra_sales_income"]),
        (JS / "pages" / "store-profit-dashboard.js",
         ["Extra Sales (non-POS)", "ytd_extra_sales_income"]),
        (JS / "pages" / "reports-pages.js",
         ["Extra Sales (non-POS", "Net Profit (gross + extra"]),
        (JS / "pages" / "extra-sales-page.js",
         ["xs-export-pdf", "xs-export-excel", "extra-sales/export"]),
    ]
    for path, needles in checks:
        src = path.read_text(encoding="utf-8")
        for needle in needles:
            assert needle in src, f"{path.name} missing {needle!r}"
